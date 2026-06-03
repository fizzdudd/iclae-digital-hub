import re
import unicodedata
from functools import wraps

from django.contrib import messages
from django.core.exceptions import ObjectDoesNotExist
from django.core.paginator import Paginator
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.db import IntegrityError, transaction
from django.db.models import Avg, Count, Max, Min, Prefetch, Q
from django.utils import timezone
from .forms import AlumnoEnlacesForm, AsignacionBadgeForm, PeriodoAcademicoForm, UsuarioCreacionForm
from .models import (
    Alumno,
    AlumnoBadge,
    AsignacionPeriodo,
    Badge,
    Bitacora,
    BitacoraEvidencia,
    CalificacionCompetencia,
    Carrera,
    CompetenciaHito,
    ConfigEvaluacionPeriodo,
    Empresa,
    EvaluacionEmpresa,
    EvaluacionHito,
    EvidenciaBitacora,
    HitoEvaluacion,
    Notificacion,
    PeriodoAcademico,
    Postulacion,
    Proyecto,
    ProyectoPeriodo,
    PuntajeCompetencia,
    RecordatorioArchivado,
    RecordatorioMasivo,
    RegistroAuditoria,
    Sede,
    TutorEmpresa,
    TutorUdd,
    Universidad,
    Usuario,
)

_LOGO_COLORS = [
    'background:linear-gradient(135deg,#bfdbfe,#93c5fd);color:#1e3a8a',
    'background:linear-gradient(135deg,#c4b5fd,#a78bfa);color:#4c1d95',
    'background:linear-gradient(135deg,#a7f3d0,#6ee7b7);color:#064e3b',
    'background:linear-gradient(135deg,#fca5a5,#f87171);color:#7f1d1d',
    'background:linear-gradient(135deg,#fed7aa,#fdba74);color:#7c2d12',
    'background:linear-gradient(135deg,#bae6fd,#7dd3fc);color:#0c4a6e',
    'background:linear-gradient(135deg,#d9f99d,#bef264);color:#365314',
]


def _empresa_logo_style(nombre):
    idx = sum(ord(c) for c in (nombre or 'E')) % len(_LOGO_COLORS)
    return _LOGO_COLORS[idx]


def _periodo_activo():
    periodo = PeriodoAcademico.objects.filter(is_active=True).first()
    if periodo:
        return periodo, periodo.nombre
    return None, 'Sin periodo activo'


def _periodo_contextual(request):
    """Período para crear/editar respetando el Modo Auditoría. Devuelve (periodo, etiqueta)."""
    user = getattr(request, 'user', None)
    es_admin = bool(user and user.is_authenticated and getattr(user, 'rol', None) == 'admin')
    if es_admin:
        audit_id = request.session.get('audit_period_id')
        if audit_id:
            periodo = PeriodoAcademico.objects.filter(pk=audit_id).first()
            if periodo:
                return periodo, periodo.nombre
    return _periodo_activo()


def _normalizar_texto(texto):
    """Cadena en minúsculas, sin acentos ni espacios extremos, para comparar nombres."""
    if texto is None:
        return ''
    descompuesto = unicodedata.normalize('NFKD', str(texto))
    sin_acentos = ''.join(c for c in descompuesto if not unicodedata.combining(c))
    return sin_acentos.strip().lower()


def _separar_nombre_completo(nombre_completo):
    """Separa un nombre completo en (nombres, apellidos) respetando apellidos compuestos."""
    # Conectores y prefijos de apellidos compuestos.
    CONECTORES = (
        'de', 'del', 'la', 'las', 'los', 'san', 'santa',
        'mac', 'mc', 'van', 'von', 'y', 'da', 'do', 'dos', 'di', 'le',
    )

    # Las comas (frecuentes al final del dato) se tratan como separadores.
    palabras = (nombre_completo or '').replace(',', ' ').split()
    if not palabras:
        return '', ''

    # Agrupa cada conector con la palabra que lo cierra (encadena consecutivos).
    bloques = []
    i = 0
    total = len(palabras)
    while i < total:
        if palabras[i].lower() in CONECTORES and i + 1 < total:
            grupo = [palabras[i]]
            i += 1
            while i < total and palabras[i].lower() in CONECTORES and i + 1 < total:
                grupo.append(palabras[i])
                i += 1
            grupo.append(palabras[i])
            i += 1
            bloques.append(' '.join(grupo))
        else:
            bloques.append(palabras[i])
            i += 1

    # Los dos últimos bloques son apellidos; el resto, nombres.
    if len(bloques) >= 3:
        nombres, apellidos = bloques[:-2], bloques[-2:]
    elif len(bloques) == 2:
        nombres, apellidos = bloques[:1], bloques[1:]
    else:
        nombres, apellidos = bloques, []

    def _titulo(partes):
        return ' '.join(palabra.capitalize() for parte in partes for palabra in parte.split())

    return _titulo(nombres), _titulo(apellidos)


# Conectores que van en minúscula dentro de un nombre propio (salvo al inicio).
_CONECTORES_TITULO = {'de', 'del', 'la', 'las', 'los', 'el', 'y', 'e', 'da', 'do', 'dos'}
# Formas jurídicas/siglas que se conservan en mayúscula en nombres de empresa.
_SIGLAS_EMPRESA = {'sa', 'spa', 'ltda', 'eirl', 'sas', 'sac', 'saic', 'saci', 'srl', 'llc', 'inc', 'ltd'}


def _titulo_es(texto, siglas=False):
    """Capitaliza como nombre propio en español.

    Cada palabra lleva inicial mayúscula y el resto en minúscula; los conectores
    (de, la, y, ...) quedan en minúscula salvo al inicio. Con siglas=True (empresas)
    se conservan en mayúscula las formas jurídicas (S.A., SpA, Ltda) y las siglas
    sin vocales (p. ej. SB, EBV). Las comas se tratan como separadores.
    """
    texto = re.sub(r'\s+', ' ', (texto or '').replace(',', ' ')).strip()
    if not texto:
        return ''

    def _capitalizar(palabra):
        return re.sub(r'[^\W\d_]+', lambda m: m.group(0)[:1].upper() + m.group(0)[1:].lower(), palabra)

    salida = []
    for indice, palabra in enumerate(texto.split()):
        nucleo = ''.join(c for c in palabra if c.isalpha())
        if not nucleo:
            salida.append(palabra)
            continue
        clave = _normalizar_texto(nucleo)
        sin_vocales = not any(v in clave for v in 'aeiou')
        if indice > 0 and clave in _CONECTORES_TITULO:
            salida.append(palabra.lower())
        elif siglas and nucleo.isupper() and ('.' in palabra or clave in _SIGLAS_EMPRESA or sin_vocales):
            salida.append(palabra.upper())
        else:
            salida.append(_capitalizar(palabra))
    return ' '.join(salida)


def _rubro_class(nombre):
    text = (nombre or '').strip().lower()
    if 'banca' in text:
        return 'banca'
    if 'fintech' in text:
        return 'fintech'
    if 'insur' in text:
        return 'insurtech'
    if 'hr' in text:
        return 'hrtech'
    if 'energ' in text:
        return 'energia'
    if 'retail' in text:
        return 'retail'
    if 'tec' in text:
        return 'tecnologia'
    return 'default'


def _build_filter_options(values):
    options = {}
    for value in values:
        text = (value or '').strip()
        if not text:
            continue
        key = text.lower()
        if key not in options:
            options[key] = text
    return [{'value': key, 'label': label} for key, label in sorted(options.items(), key=lambda item: item[1].lower())]


def _bitacora_estado_label(estado):
    estado_text = (estado or 'pendiente').strip().lower()
    return {
        'aprobado': 'Aprobado',
        'aprobada': 'Aprobado',
        'pendiente': 'Pendiente',
        'corregido': 'Corregido',
        'reprobado': 'Reprobado',
        'enviada': 'Enviada',
        'enviado': 'Enviada',
        'en_curso': 'En curso',
    }.get(estado_text, estado_text.title())


def _get_user_rol(request):
    """Rol del usuario autenticado. Un admin puede previsualizar otros roles vía ?rol= (demo)."""
    user = getattr(request, 'user', None)
    real_rol = getattr(user, 'rol', None) if (user is not None and user.is_authenticated) else None

    if real_rol == 'admin':
        rol_param = (request.GET.get('rol') or '').strip().lower()
        if rol_param in ('admin', 'alumno', 'tutor_udd', 'tutor_empresa'):
            request.session['rol_preview'] = rol_param
        preview = request.session.get('rol_preview')
        if preview in ('admin', 'alumno', 'tutor_udd', 'tutor_empresa'):
            return preview

    return real_rol or 'admin'


def require_roles(*roles):
    """Restringe la vista a ciertos roles (evalúa el rol real, no la previsualización)."""
    allowed = set(roles)

    def decorator(view_func):
        @wraps(view_func)
        def _wrapped(request, *args, **kwargs):
            rol = getattr(getattr(request, 'user', None), 'rol', None)
            if rol == 'admin' or rol in allowed:
                return view_func(request, *args, **kwargs)
            messages.error(request, 'No tienes permisos para acceder a esta sección.')
            return redirect('dashboard')
        return _wrapped
    return decorator


def _fmt_bytes(n):
    """Formatea un tamaño en bytes a KB/MB legible (cadena vacía si no hay dato)."""
    if not n:
        return ''
    if n >= 1024 * 1024:
        return f"{round(n / (1024 * 1024), 1)} MB"
    return f"{round(n / 1024, 1)} KB"


def _ext_de(nombre):
    import os
    return os.path.splitext(nombre or '')[1].lstrip('.').lower()


def _bitacora_estado_class(estado):
    estado_text = (estado or 'pendiente').strip().lower()
    if estado_text in {'aprobado', 'aprobada'}:
        return 'aprobado'
    if estado_text in {'corregido', 'corregida'}:
        return 'corregido'
    if estado_text in {'reprobado', 'reprobada'}:
        return 'reprobado'
    if estado_text in {'enviada', 'enviado'}:
        return 'enviado'
    return 'pendiente'


def calcular_nota_bitacoras(semanas_aprobadas, total_semanas, porcentaje_exigencia=60):
    """Nota de bitácoras (1.0–7.0) según cumplimiento de semanas y exigencia del admin."""
    if not total_semanas or total_semanas <= 0:
        return 1.0

    cumplimiento = semanas_aprobadas / total_semanas

    exigencia = (porcentaje_exigencia or 0) / 100.0
    # Salvaguardas para evitar divisiones por cero con exigencias extremas.
    if exigencia <= 0:
        exigencia = 0.6
    if exigencia >= 1:
        exigencia = 0.99

    if cumplimiento < exigencia:
        nota = 1.0 + 3.0 * (cumplimiento / exigencia)
    else:
        nota = 4.0 + 3.0 * ((cumplimiento - exigencia) / (1.0 - exigencia))

    return round(nota, 1)


def _render_buscador_alumno(request, periodo, periodo_label, user_rol, titulo, accion_label):
    """Lista filtrable de alumnos (nombre/correo y carrera) para que el admin elija uno."""
    q = (request.GET.get('q') or '').strip()
    carrera_filter = (request.GET.get('carrera') or '').strip().lower()

    pps = (
        ProyectoPeriodo.objects.filter(periodo=periodo, alumno__isnull=False)
        .select_related('alumno__id', 'alumno__carrera', 'proyecto__empresa')
        .order_by('alumno__id__nombre', 'alumno__id__apellido')
    )

    carreras = {}
    alumnos = []
    for pp in pps:
        u = pp.alumno.id
        nombre = f"{u.nombre} {u.apellido}".strip()
        email = u.email or ''
        carrera_nombre = pp.alumno.carrera.nombre if pp.alumno.carrera else ''
        if carrera_nombre:
            carreras[carrera_nombre.lower()] = carrera_nombre
        if q and q.lower() not in nombre.lower() and q.lower() not in email.lower():
            continue
        if carrera_filter and carrera_nombre.lower() != carrera_filter:
            continue
        alumnos.append({
            'id': str(pp.alumno.id_id),
            'nombre': nombre,
            'email': email,
            'iniciales': (''.join(p[0] for p in nombre.split()[:2]).upper() or 'AL'),
            'carrera': carrera_nombre or 'Sin carrera',
            'empresa': pp.proyecto.empresa.nombre if pp.proyecto and pp.proyecto.empresa else 'Sin empresa',
            'proyecto': pp.proyecto.titulo if pp.proyecto else 'Sin proyecto',
        })

    carrera_options = [
        {'value': key, 'label': label}
        for key, label in sorted(carreras.items(), key=lambda item: item[1].lower())
    ]

    return render(request, 'pages/buscador_alumno.html', {
        'periodo_label': periodo_label,
        'user_rol': user_rol,
        'buscador_titulo': titulo,
        'accion_label': accion_label,
        'alumnos': alumnos,
        'alumnos_count': len(alumnos),
        'carrera_options': carrera_options,
        'filters': {'q': q, 'carrera': carrera_filter},
    })


def _calcular_boletin(periodo, current_pp):
    """Boletín del alumno: devuelve (rows, promedio_final, en_calculo, pendientes)."""
    total_semanas = periodo.total_semanas or 13
    config = ConfigEvaluacionPeriodo.objects.filter(periodo=periodo).first()
    peso_bitacoras = config.peso_bitacoras_pct if config else 0
    exigencia = config.umbral_bitacoras_pct if (config and config.umbral_bitacoras_pct) else 60

    aprobadas = 0
    for b in Bitacora.objects.filter(proyecto_periodo=current_pp):
        if (_bitacora_estado_class(b.estado_emp) == 'aprobado'
                and _bitacora_estado_class(b.estado_udd) == 'aprobado'):
            aprobadas += 1
    nota_bitacoras = calcular_nota_bitacoras(aprobadas, total_semanas, exigencia)

    rows = [{
        'nombre': 'Bitácoras',
        'detalle': '%s de %s semanas aprobadas' % (aprobadas, total_semanas),
        'peso': peso_bitacoras,
        'nota': nota_bitacoras,
        'evaluado': True,
        'auto': True,
    }]

    suma_pond = nota_bitacoras * peso_bitacoras
    suma_pesos = peso_bitacoras
    pendientes = 0

    eval_map = {e.hito_id: e for e in EvaluacionHito.objects.filter(proyecto_periodo=current_pp)}
    for h in HitoEvaluacion.objects.filter(periodo=periodo).order_by('orden', 'semana'):
        ev = eval_map.get(h.pk)
        nota = float(ev.nota_calculada) if (ev and ev.nota_calculada is not None) else None
        evaluado = nota is not None
        if evaluado:
            suma_pond += nota * (h.peso_pct or 0)
            suma_pesos += (h.peso_pct or 0)
        else:
            pendientes += 1
        rows.append({
            'nombre': h.nombre,
            'detalle': 'Semana %s' % h.semana,
            'peso': h.peso_pct or 0,
            'nota': nota,
            'evaluado': evaluado,
            'auto': False,
        })

    en_calculo = pendientes > 0
    promedio_final = round(suma_pond / suma_pesos, 1) if suma_pesos else None
    return rows, promedio_final, en_calculo, pendientes


def _render_dashboard_tutor(request, periodo, periodo_label, user_rol, tutor_tipo, project_periods):
    """Alumnos asignados al tutor con su avance y semanas pendientes de su revisión."""
    total_weeks = periodo.total_semanas or 13
    pps = list(project_periods)
    pp_ids = [pp.pk for pp in pps]

    aprob_por_pp = {}
    pend_por_pp = {}
    if pp_ids:
        for b in Bitacora.objects.filter(proyecto_periodo_id__in=pp_ids):
            ce = _bitacora_estado_class(b.estado_emp)
            cu = _bitacora_estado_class(b.estado_udd)
            if ce == 'aprobado' and cu == 'aprobado':
                aprob_por_pp[b.proyecto_periodo_id] = aprob_por_pp.get(b.proyecto_periodo_id, 0) + 1
            propio = cu if tutor_tipo == 'udd' else ce
            if b.fecha_envio and propio in ('enviado', 'pendiente'):
                pend_por_pp[b.proyecto_periodo_id] = pend_por_pp.get(b.proyecto_periodo_id, 0) + 1

    rows = []
    for pp in pps:
        aprobadas = aprob_por_pp.get(pp.pk, 0)
        rows.append({
            'alumno_id': str(pp.alumno.id_id),
            'alumno_nombre': f"{pp.alumno.id.nombre} {pp.alumno.id.apellido}".strip(),
            'empresa': pp.proyecto.empresa.nombre if pp.proyecto and pp.proyecto.empresa else 'Sin empresa',
            'proyecto': pp.proyecto.titulo if pp.proyecto else 'Sin proyecto',
            'aprobadas': aprobadas,
            'total': total_weeks,
            'pct': int((aprobadas / max(total_weeks, 1)) * 100),
            'pendientes': pend_por_pp.get(pp.pk, 0),
        })

    rows.sort(key=lambda r: (-r['pendientes'], r['alumno_nombre']))
    pendientes_total = sum(r['pendientes'] for r in rows)

    return render(request, 'pages/dashboard_tutor.html', {
        'periodo_label': periodo_label,
        'user_rol': user_rol,
        'tutor_tipo': tutor_tipo,
        'rows': rows,
        'total_alumnos': len(rows),
        'pendientes_total': pendientes_total,
    })


@require_roles('alumno', 'tutor_udd', 'tutor_empresa')
def bitacora_view(request):
    user_rol = _get_user_rol(request)
    # Aislamiento de período respetando el Modo Auditoría del administrador.
    periodo, periodo_label = _periodo_contextual(request)

    if not periodo:
        return render(
            request,
            'pages/bitacora.html',
            {
                'periodo_label': periodo_label,
                'user_rol': user_rol,
                'bitacora_empty': True,
                'message': 'No hay un periodo activo para mostrar la bitácora del alumno.',
            },
        )

    user = request.user
    real_rol = getattr(user, 'rol', None)
    # El rol efectivo respeta la previsualización del admin (demo); el rol real
    # se usa para el alcance de datos por seguridad.
    is_tutor = user_rol in ('tutor_udd', 'tutor_empresa')
    tutor_tipo = 'udd' if user_rol == 'tutor_udd' else ('empresa' if user_rol == 'tutor_empresa' else None)

    alumno_uid = (request.GET.get('alumno') or '').strip()
    semana_query = request.GET.get('sem')

    project_periods = (
        ProyectoPeriodo.objects.filter(periodo=periodo, alumno__isnull=False)
        .select_related('proyecto__empresa', 'alumno__id', 'sede')
        .prefetch_related('tutores_udd__id', 'tutores_empresa__id')
        .order_by('alumno__id__nombre', 'alumno__id__apellido', 'proyecto__titulo')
    )

    # ── Alcance por rol (seguridad, según el rol REAL) ──
    # Alumno: solo su propia bitácora. Tutor: solo sus alumnos asignados.
    # Admin: todos (incluido cuando previsualiza otro rol).
    if real_rol == 'alumno':
        project_periods = project_periods.filter(alumno_id=user.id)
        alumno_uid = str(user.id)
    elif real_rol == 'tutor_udd':
        project_periods = project_periods.filter(tutores_udd=user.id)
    elif real_rol == 'tutor_empresa':
        project_periods = project_periods.filter(tutores_empresa=user.id)

    # ── Bifurcación por rol ──
    # Un tutor que entra sin un alumno seleccionado ve su panel (dashboard) con
    # la lista de alumnos asignados. Al elegir un alumno (?alumno=) ve el detalle.
    if is_tutor and not alumno_uid and request.method == 'GET':
        return _render_dashboard_tutor(
            request, periodo, periodo_label, user_rol, tutor_tipo, project_periods
        )

    # ── Administrador sin alumno seleccionado: panel de búsqueda ──
    # Al elegir un alumno (?alumno=) se renderiza el detalle de su bitácora.
    is_admin = user_rol == 'admin'
    if is_admin and not alumno_uid and request.method == 'GET':
        return _render_buscador_alumno(
            request, periodo, periodo_label, user_rol,
            'Bitácoras · Selección de alumno', 'Ver bitácora',
        )

    current_pp = None
    if alumno_uid:
        current_pp = project_periods.filter(alumno__id_id=alumno_uid).first()
    if current_pp is None:
        current_pp = project_periods.first()

    if request.method == 'POST' and current_pp:
        semana_post = request.POST.get('semana') or semana_query or (current_pp.semana_actual or 1)
        try:
            semana_post_int = int(semana_post)
        except (TypeError, ValueError):
            semana_post_int = current_pp.semana_actual or 1

        texto = (request.POST.get('texto') or '').strip()
        action = (request.POST.get('action') or 'draft').strip().lower()

        # ── Revisión del tutor (Aprobar / Por corregir / Reprobar) ──
        if action == 'review' and is_tutor:
            destino_review = f"{request.path}?alumno={current_pp.alumno.id_id}&sem={semana_post_int}"
            decision = (request.POST.get('decision') or '').strip().lower()
            feedback = (request.POST.get('feedback') or '').strip()
            mapping = {'aprobar': 'aprobada', 'corregir': 'corregida', 'reprobar': 'reprobada'}
            nuevo_estado = mapping.get(decision)
            if not nuevo_estado:
                messages.error(request, 'Selecciona una decisión válida para la evaluación.')
                return redirect(destino_review)
            if nuevo_estado in ('corregida', 'reprobada') and not feedback:
                messages.error(request, 'Debes escribir un comentario para "Por corregir" o "Reprobar".')
                return redirect(destino_review)

            bitacora_rev = Bitacora.objects.filter(
                proyecto_periodo=current_pp, semana=semana_post_int
            ).first()
            if not bitacora_rev or not bitacora_rev.fecha_envio:
                messages.error(request, 'No se puede evaluar una semana que aún no ha sido enviada.')
                return redirect(destino_review)

            ahora = timezone.now()
            if tutor_tipo == 'udd':
                bitacora_rev.estado_udd = nuevo_estado
                bitacora_rev.feedback_udd = feedback
                bitacora_rev.fecha_revision_udd = ahora
            else:
                bitacora_rev.estado_emp = nuevo_estado
                bitacora_rev.feedback_emp = feedback
                bitacora_rev.fecha_revision_emp = ahora
            bitacora_rev.updated_at = ahora
            bitacora_rev.save()

            if bitacora_rev.esta_cerrada:
                messages.success(request, 'Evaluación guardada. La bitácora quedó cerrada (ambos tutores aprobaron).')
            else:
                messages.success(request, 'Evaluación guardada correctamente.')
            return redirect(destino_review)

        # A partir de aquí, solo el alumno (o admin) edita/envía el registro.

        # Si un tutor ya aprobó la semana, la bitácora queda bloqueada: no se
        # permite editar ni reenviar (ni desde el front ni vía POST directo).
        existing = Bitacora.objects.filter(proyecto_periodo=current_pp, semana=semana_post_int).first()
        if existing and (
            _bitacora_estado_class(existing.estado_emp) == 'aprobado'
            or _bitacora_estado_class(existing.estado_udd) == 'aprobado'
        ):
            messages.error(request, 'Esta bitácora ya fue aprobada por un tutor y no puede editarse.')
            return redirect(f"{request.path}?alumno={current_pp.alumno.id_id}&sem={semana_post_int}")

        bitacora, _ = Bitacora.objects.get_or_create(
            proyecto_periodo=current_pp,
            semana=semana_post_int,
            defaults={'texto': texto, 'created_at': timezone.now()},
        )
        bitacora.texto = texto
        bitacora.updated_at = timezone.now()
        if action == 'send':
            bitacora.fecha_envio = timezone.now()
            bitacora.estado_emp = bitacora.estado_emp or 'pendiente'
            bitacora.estado_udd = bitacora.estado_udd or 'pendiente'
        bitacora.save()
        return redirect(f"{request.path}?alumno={current_pp.alumno.id_id}&sem={semana_post_int}")

    if current_pp:
        alumno = current_pp.alumno
        alumno_name = f"{alumno.id.nombre} {alumno.id.apellido}".strip()
        company_name = current_pp.proyecto.empresa.nombre if current_pp.proyecto and current_pp.proyecto.empresa else 'Sin empresa'
        project_name = current_pp.proyecto.titulo if current_pp.proyecto else 'Sin proyecto'
        total_weeks = periodo.total_semanas or 13
        selected_week = int(semana_query) if semana_query and str(semana_query).isdigit() else (current_pp.semana_actual or 1)
        selected_week = max(1, min(selected_week, total_weeks))

        bitacoras_qs = Bitacora.objects.filter(proyecto_periodo=current_pp).prefetch_related('bitacoraevidencia_set', 'evidencias')
        bitacoras_by_week = {item.semana: item for item in bitacoras_qs}

        # Un tutor sin semana explícita aterriza en la más antigua pendiente de SU
        # revisión (enviada y sin resolver por su caja). Si no hay, deja la actual.
        if is_tutor and not (semana_query and str(semana_query).isdigit()):
            for n in range(1, total_weeks + 1):
                b = bitacoras_by_week.get(n)
                if not (b and b.fecha_envio):
                    continue
                propio = _bitacora_estado_class(b.estado_udd if tutor_tipo == 'udd' else b.estado_emp)
                if propio in ('enviado', 'pendiente'):
                    selected_week = n
                    break

        weeks = []
        approved = corrected = pending = en_curso = 0
        for week_number in range(1, total_weeks + 1):
            bitacora = bitacoras_by_week.get(week_number)
            if bitacora:
                estado_emp = _bitacora_estado_class(bitacora.estado_emp)
                estado_udd = _bitacora_estado_class(bitacora.estado_udd)
                has_sent = bool(bitacora.fecha_envio)
                if estado_emp == 'aprobado' and estado_udd == 'aprobado':
                    progress_state = 'aprobado'
                    approved += 1
                elif estado_emp == 'corregido' or estado_udd == 'corregido':
                    progress_state = 'corregido'
                    corrected += 1
                elif estado_emp == 'reprobado' or estado_udd == 'reprobado':
                    progress_state = 'reprobado'
                elif has_sent:
                    # Enviada y a la espera de revisión de los tutores.
                    progress_state = 'en_curso'
                    en_curso += 1
                else:
                    progress_state = 'pendiente'
                    pending += 1
                texto_bitacora = bitacora.texto or ''
                fecha_envio = bitacora.fecha_envio
                # Evidencias: combinamos el legado (BitacoraEvidencia, solo URL) con
                # las subidas nuevas (EvidenciaBitacora, archivo real).
                evidencias = [
                    {
                        'nombre': ev.nombre_archivo,
                        'url': ev.url,
                        'tipo': (ev.tipo_archivo or _ext_de(ev.nombre_archivo) or 'archivo').lower(),
                        'tamano': _fmt_bytes(ev.tamaño_bytes),
                    }
                    for ev in bitacora.bitacoraevidencia_set.all()
                ]
                for ev in bitacora.evidencias.all():
                    try:
                        size = ev.archivo.size if ev.archivo else 0
                    except (OSError, ValueError):
                        size = 0
                    evidencias.append({
                        'nombre': ev.nombre,
                        'url': ev.archivo.url if ev.archivo else '',
                        'tipo': ev.extension or 'archivo',
                        'tamano': _fmt_bytes(size),
                    })
            else:
                progress_state = 'aprobado' if week_number < (current_pp.semana_actual or 1) and week_number in bitacoras_by_week else 'pendiente'
                if week_number >= (current_pp.semana_actual or 1):
                    pending += 1
                texto_bitacora = ''
                fecha_envio = None
                has_sent = False
                evidencias = []

            weeks.append(
                {
                    'n': week_number,
                    'selected': week_number == selected_week,
                    'current': week_number == (current_pp.semana_actual or 1),
                    'future': week_number > (current_pp.semana_actual or 1),
                    'status': progress_state,
                    'status_label': _bitacora_estado_label(progress_state),
                    'has_sent': has_sent,
                    'texto': texto_bitacora,
                    'fecha_envio': fecha_envio,
                    'evidencias': evidencias,
                    'estado_emp': _bitacora_estado_class(bitacora.estado_emp) if bitacora else 'pendiente',
                    'estado_udd': _bitacora_estado_class(bitacora.estado_udd) if bitacora else 'pendiente',
                }
            )

        selected_week_data = next((item for item in weeks if item['n'] == selected_week), weeks[0])
        selected_bitacora = bitacoras_by_week.get(selected_week)
        selected_text = selected_bitacora.texto if selected_bitacora and selected_bitacora.texto is not None else selected_week_data['texto']
        selected_evidencias = selected_week_data['evidencias']
        locked_by_tutor = bool(selected_bitacora) and (
            _bitacora_estado_class(selected_bitacora.estado_emp) == 'aprobado'
            or _bitacora_estado_class(selected_bitacora.estado_udd) == 'aprobado'
        )
        editable = (
            selected_week == (current_pp.semana_actual or 1)
            and not selected_week_data['future']
            and not locked_by_tutor
        )
        progress_total = max(total_weeks, 1)
        progress_pct = int((approved / progress_total) * 100)

        # Tutor info for selected week (hasta 2 tutores por tipo).
        emp_users = [t.id for t in current_pp.tutores_empresa.all()]
        udd_users = [t.id for t in current_pp.tutores_udd.all()]
        tutor_empresa_nombre = ', '.join(f"{u.nombre} {u.apellido}".strip() for u in emp_users) or 'Sin tutor empresa'
        tutor_udd_nombre = ', '.join(f"{u.nombre} {u.apellido}".strip() for u in udd_users) or 'Sin tutor UDD'
        tutor_empresa_initials = (''.join(p[0] for p in f"{emp_users[0].nombre} {emp_users[0].apellido}".split()[:2]).upper() or 'TE') if emp_users else 'TE'
        tutor_udd_initials = (''.join(p[0] for p in f"{udd_users[0].nombre} {udd_users[0].apellido}".split()[:2]).upper() or 'TU') if udd_users else 'TU'

        selected_estado_emp = _bitacora_estado_label(selected_bitacora.estado_emp if selected_bitacora else None)
        selected_estado_udd = _bitacora_estado_label(selected_bitacora.estado_udd if selected_bitacora else None)
        selected_estado_emp_class = _bitacora_estado_class(selected_bitacora.estado_emp if selected_bitacora else None)
        selected_estado_udd_class = _bitacora_estado_class(selected_bitacora.estado_udd if selected_bitacora else None)
        selected_feedback_emp = (selected_bitacora.feedback_emp or '') if selected_bitacora else ''
        selected_feedback_udd = (selected_bitacora.feedback_udd or '') if selected_bitacora else ''

        # Datos del veredicto de la caja del tutor ACTIVO (la que puede editar).
        # review_ya_evaluo indica si ya guardó una decisión (aprobado/corregido/
        # reprobado) para mostrar el modo lectura en lugar del formulario.
        if tutor_tipo == 'udd':
            review_estado_class = selected_estado_udd_class
            review_estado_label = selected_estado_udd
            review_feedback = selected_feedback_udd
        else:
            review_estado_class = selected_estado_emp_class
            review_estado_label = selected_estado_emp
            review_feedback = selected_feedback_emp
        review_ya_evaluo = review_estado_class in ('aprobado', 'corregido', 'reprobado')

        context = {
            'periodo_label': periodo_label,
            'student_name': alumno_name,
            'student_company': company_name,
            'project_name': project_name,
            'selected_week': selected_week,
            'current_week': current_pp.semana_actual or 1,
            'weeks_total': total_weeks,
            'weeks': weeks,
            'selected_text': selected_text,
            'selected_evidences': selected_evidencias,
            'editable': editable,
            'progress_pct': progress_pct,
            'approved_count': approved,
            'corrected_count': corrected,
            'en_curso_count': en_curso,
            'pending_count': pending,
            'selected_week_label': selected_week_data['status_label'],
            'selected_week_status': selected_week_data['status'],
            'selected_week_has_sent': selected_week_data['has_sent'],
            'selected_week_fecha': selected_week_data['fecha_envio'],
            'selected_week_future': selected_week_data['future'],
            'can_choose_weeks': True,
            'is_tutor': is_tutor,
            'tutor_tipo': tutor_tipo,
            'can_edit_registro': editable and not is_tutor,
            'can_review': is_tutor and selected_week_data['has_sent'],
            'review_estado_class': review_estado_class,
            'review_estado_label': review_estado_label,
            'review_feedback': review_feedback,
            'review_ya_evaluo': review_ya_evaluo,
            'project_period': current_pp,
            'student_options': [
                {
                    'id': str(item.alumno.id_id),
                    'label': f"{item.alumno.id.nombre} {item.alumno.id.apellido}".strip(),
                    'company': item.proyecto.empresa.nombre if item.proyecto and item.proyecto.empresa else '',
                }
                for item in project_periods
            ],
            'selected_student_id': str(current_pp.alumno.id_id),
            'bitacora_weekday_hint': 'Sé concreto: qué hiciste, con quién y qué resultado tuvo.',
            'tutor_empresa_nombre': tutor_empresa_nombre,
            'tutor_empresa_initials': tutor_empresa_initials,
            'tutor_udd_nombre': tutor_udd_nombre,
            'tutor_udd_initials': tutor_udd_initials,
            'selected_estado_emp': selected_estado_emp,
            'selected_estado_udd': selected_estado_udd,
            'selected_estado_emp_class': selected_estado_emp_class,
            'selected_estado_udd_class': selected_estado_udd_class,
            'locked_by_tutor': locked_by_tutor,
            'selected_feedback_emp': selected_feedback_emp,
            'selected_feedback_udd': selected_feedback_udd,
            'user_rol': user_rol,
            'volver_url': request.path if is_admin else '',
        }
        return render(request, 'pages/bitacora.html', context)

    return render(
        request,
        'pages/bitacora.html',
        {
            'periodo_label': periodo_label,
            'user_rol': user_rol,
            'bitacora_empty': True,
            'message': 'No hay alumnos asignados a un proyecto en el periodo activo.',
        },
    )


def dashboard_view(request):
    user_rol = _get_user_rol(request)
    # Handle POST for sending reminder
    if request.method == 'POST' and request.POST.get('action') == 'send_reminder':
        segmento = (request.POST.get('segmento') or 'todos').strip()[:50]
        mensaje = (request.POST.get('mensaje') or '').strip()
        try:
            cantidad = int(request.POST.get('cantidad') or 0)
        except (TypeError, ValueError):
            cantidad = 0
        if mensaje:
            admin_user = Usuario.objects.filter(is_active=True).order_by('created_at').first()
            if admin_user:
                try:
                    RecordatorioMasivo.objects.create(
                        enviado_por=admin_user,
                        segmento=segmento or 'todos',
                        mensaje=mensaje,
                        cantidad_envios=cantidad,
                        fecha_envio=timezone.now(),
                    )
                except Exception:
                    pass
        return redirect('dashboard')

    # KPIs base
    alumnos_count = Alumno.objects.count()
    total_alumnos = alumnos_count or 1
    empresas_count = Empresa.objects.count()
    practicas_activas = ProyectoPeriodo.objects.filter(estado='en_curso').count()
    tutores_empresa_count = TutorEmpresa.objects.count()
    tutores_udd_count = TutorUdd.objects.count()
    # Aislamiento de período respetando el Modo Auditoría del administrador.
    periodo_activo, periodo_label = _periodo_contextual(request)

    # Funnel (unique alumnos to avoid counting multiple applications per student)
    postulantes_count = Postulacion.objects.values('alumno').distinct().count()
    seleccionados_count = (
        Postulacion.objects.filter(
            Q(estado__icontains='seleccion') | Q(estado__icontains='acept')
        )
        .values('alumno')
        .distinct()
        .count()
    )
    contratados_count = ProyectoPeriodo.objects.filter(estado__icontains='contrat').count()

    # Top empresas por practicantes
    top_empresas_qs = list(
        Empresa.objects.annotate(
            total_practicantes=Count(
                'proyecto__proyectoperiodo',
                filter=Q(
                    proyecto__proyectoperiodo__alumno__isnull=False,
                    proyecto__proyectoperiodo__periodo=periodo_activo,
                ),
            ),
            total_proyectos=Count('proyecto', distinct=True),
        ).order_by('-total_practicantes')[:6]
    )
    max_practicantes = max((e.total_practicantes for e in top_empresas_qs), default=1) or 1
    top_empresas = [
        {
            'rank': idx,
            'nombre': e.nombre,
            'rubro': e.rubro or 'Sin rubro',
            'rubro_class': _rubro_class(e.rubro),
            'total_practicantes': e.total_practicantes,
            'total_proyectos': e.total_proyectos,
            'pct': int((e.total_practicantes / max_practicantes) * 100),
        }
        for idx, e in enumerate(top_empresas_qs, start=1)
    ]

    # Rendimiento académico
    notas = ProyectoPeriodo.objects.exclude(nota_final__isnull=True)
    notas_stats = notas.aggregate(
        promedio=Avg('nota_final'),
        maxima=Max('nota_final'),
        minima=Min('nota_final'),
    )

    top_competencias_qs = (
        PuntajeCompetencia.objects.values('competencia__nombre')
        .annotate(promedio=Avg('nota'))
        .order_by('-promedio')[:5]
    )
    top_competencias = []
    for idx, comp in enumerate(top_competencias_qs, start=1):
        promedio = float(comp['promedio'] or 0)
        pct = max(0, min(100, int((promedio / 7.0) * 100)))
        nivel = 'Alto' if promedio >= 6 else 'Bajo' if promedio < 5 else 'Medio'
        top_competencias.append(
            {
                'rank': idx,
                'nombre': comp['competencia__nombre'] or 'Competencia',
                'score': round(promedio, 1),
                'pct': pct,
                'nivel': nivel,
                'nivel_class': 'green' if nivel == 'Alto' else 'rose' if nivel == 'Bajo' else 'blue',
                'nivel_arrow': '↑ ' if nivel == 'Alto' else '↓ ' if nivel == 'Bajo' else '',
            }
        )

    # Distribuciones por sede (para gráfico doughnut)
    sedes_qs = (
        ProyectoPeriodo.objects.values('sede__nombre')
        .annotate(total=Count('pk'))
        .order_by('-total')[:6]
    )
    sedes_labels = [row['sede__nombre'] or 'Sin sede' for row in sedes_qs]
    sedes_values = [row['total'] for row in sedes_qs]

    # Distribución por rubro (all rubros, no limit, so totals match empresas_count KPI)
    rubros_qs = list(
        Empresa.objects.values('rubro')
        .annotate(total=Count('pk'))
        .order_by('-total')
    )
    rubro_max = max([row['total'] for row in rubros_qs], default=1) or 1
    bar_colors_rubro = ['bar-blue', 'bar-green', 'bar-violet', 'bar-amber', 'bar-rose', 'bar-cyan', 'bar-orange', 'bar-blue']
    rubros = [
        {
            'nombre': row['rubro'] or 'Sin rubro',
            'total': row['total'],
            'pct': int((row['total'] / rubro_max) * 100),
            'bar_class': bar_colors_rubro[i % len(bar_colors_rubro)],
        }
        for i, row in enumerate(rubros_qs)
    ]

    # Modalidad
    modalidad_qs = Proyecto.objects.values('modalidad').annotate(total=Count('pk')).order_by('-total')
    modalidad_map = {'presencial': 0, 'hibrido': 0, 'híbrido': 0, 'remoto': 0}
    for row in modalidad_qs:
        key = (row['modalidad'] or '').strip().lower()
        if key in modalidad_map:
            modalidad_map[key] += row['total']
    modalidad_total = sum(modalidad_map.values()) or 1
    modalidad_data = [
        {
            'label': 'Presencial',
            'count': modalidad_map['presencial'],
            'pct': int((modalidad_map['presencial'] / modalidad_total) * 100),
            'class': 'presencial',
        },
        {
            'label': 'Híbrido',
            'count': modalidad_map['hibrido'] + modalidad_map['híbrido'],
            'pct': int(((modalidad_map['hibrido'] + modalidad_map['híbrido']) / modalidad_total) * 100),
            'class': 'hibrido',
        },
        {
            'label': 'Remoto',
            'count': modalidad_map['remoto'],
            'pct': int((modalidad_map['remoto'] / modalidad_total) * 100),
            'class': 'remoto',
        },
    ]

    # Distribución por carrera
    carreras_qs = (
        Alumno.objects.values('carrera__nombre')
        .annotate(total=Count('pk'))
        .order_by('-total')[:6]
    )
    carrera_max = max([row['total'] for row in carreras_qs], default=1) or 1
    bar_colors_carrera = ['bar-blue', 'bar-green', 'bar-violet', 'bar-rose', 'bar-amber', 'bar-cyan']
    carreras_dist = [
        {
            'nombre': row['carrera__nombre'] or 'Sin carrera',
            'total': row['total'],
            'pct': int((row['total'] / carrera_max) * 100),
            'bar_class': bar_colors_carrera[i % len(bar_colors_carrera)],
        }
        for i, row in enumerate(carreras_qs)
    ]

    # Distribución por sede/región
    region_qs = (
        ProyectoPeriodo.objects.values('sede__nombre')
        .annotate(total=Count('pk'))
        .order_by('-total')[:5]
    )
    region_total = sum(row['total'] for row in region_qs) or 1
    region_colors = ['#3b82f6', '#8b5cf6', '#06b6d4', '#22c55e', '#f59e0b']
    region_bar_classes = ['bar-blue', 'bar-violet', 'bar-cyan', 'bar-green', 'bar-amber']
    region_dist = [
        {
            'nombre': row['sede__nombre'] or 'Sin sede',
            'total': row['total'],
            'pct': int((row['total'] / region_total) * 100),
            'dot_color': region_colors[i % len(region_colors)],
            'bar_class': region_bar_classes[i % len(region_bar_classes)],
        }
        for i, row in enumerate(region_qs)
    ]

    # eNPS promedio empresa (único campo disponible en la BD)
    enps_empresa_avg = round(
        float(
            Empresa.objects.exclude(enps_score__isnull=True)
            .aggregate(avg=Avg('enps_score'))['avg'] or 0
        ),
        1,
    )

    # ── Empleabilidad digital: solo alumnos activos del período contextual ──
    if periodo_activo is not None:
        alumnos_activos_qs = Alumno.objects.filter(proyectoperiodo__periodo=periodo_activo).distinct()
    else:
        alumnos_activos_qs = Alumno.objects.none()
    alumnos_activos_count = alumnos_activos_qs.count()
    base_activos = alumnos_activos_count or 1

    # Totales por enlace (alumno con el campo correspondiente lleno).
    linkedin_completo = alumnos_activos_qs.exclude(url_linkedin__isnull=True).exclude(url_linkedin='').count()
    cv_completo = alumnos_activos_qs.exclude(url_cv__isnull=True).exclude(url_cv='').count()
    video_completo = alumnos_activos_qs.exclude(url_youtube__isnull=True).exclude(url_youtube='').count()
    # Perfil 100% completo: los tres enlaces llenos a la vez.
    profile_completo_count = (
        alumnos_activos_qs.exclude(url_linkedin__isnull=True)
        .exclude(url_linkedin='')
        .exclude(url_cv__isnull=True)
        .exclude(url_cv='')
        .exclude(url_youtube__isnull=True)
        .exclude(url_youtube='')
        .count()
    )

    # Porcentajes exactos respecto a los alumnos activos.
    linkedin_pct = int((linkedin_completo / base_activos) * 100)
    cv_pct = int((cv_completo / base_activos) * 100)
    video_pct = int((video_completo / base_activos) * 100)
    profile_completo_pct = int((profile_completo_count / base_activos) * 100)
    sin_perfil_completo = alumnos_activos_count - profile_completo_count

    profile_data = [
        {
            'label': 'LinkedIn',
            'count': linkedin_completo,
            'total': alumnos_activos_count,
            'pct': linkedin_pct,
            'stroke': '#2563eb',
            'offset': round(188.5 - (188.5 * linkedin_pct / 100), 1),
        },
        {
            'label': 'CV subido',
            'count': cv_completo,
            'total': alumnos_activos_count,
            'pct': cv_pct,
            'stroke': '#16a34a',
            'offset': round(188.5 - (188.5 * cv_pct / 100), 1),
        },
        {
            'label': 'Video YouTube',
            'count': video_completo,
            'total': alumnos_activos_count,
            'pct': video_pct,
            'stroke': '#dc2626',
            'offset': round(188.5 - (188.5 * video_pct / 100), 1),
        },
    ]

    # Badges
    badge_data = list(
        AlumnoBadge.objects.values('badge__nombre', 'badge__descripcion', 'badge__icono')
        .annotate(total=Count('pk'))
        .order_by('-total')[:4]
    )
    badge_total = AlumnoBadge.objects.count()
    badge_tipos = AlumnoBadge.objects.values('badge').distinct().count()
    alumnos_con_badge = AlumnoBadge.objects.values('alumno').distinct().count()
    alumnos_sin_badge = alumnos_count - alumnos_con_badge

    def funnel_pct(value):
        base = postulantes_count or 1
        return round((value / base) * 100, 1)

    context = {
        'periodo_label': periodo_label,
        'alumnos_count': alumnos_count,
        'empresas_count': empresas_count,
        'practicas_activas': practicas_activas,
        'tutores_empresa_count': tutores_empresa_count,
        'tutores_udd_count': tutores_udd_count,
        'postulantes_count': postulantes_count,
        'seleccionados_count': seleccionados_count,
        'contratados_count': contratados_count,
        'postulantes_pct': funnel_pct(postulantes_count),
        'seleccionados_pct': funnel_pct(seleccionados_count),
        'en_curso_pct': funnel_pct(practicas_activas),
        'contratados_pct': funnel_pct(contratados_count),
        'nota_promedio': round(float(notas_stats['promedio'] or 0), 1),
        'nota_max': round(float(notas_stats['maxima'] or 0), 1),
        'nota_min': round(float(notas_stats['minima'] or 0), 1),
        'top_competencias': top_competencias,
        'top_empresas': top_empresas,
        'sedes_labels': sedes_labels,
        'sedes_values': sedes_values,
        'rubros': rubros,
        'modalidad_data': modalidad_data,
        'profile_data': profile_data,
        'alumnos_activos_count': alumnos_activos_count,
        'linkedin_completo': linkedin_completo,
        'cv_completo': cv_completo,
        'video_completo': video_completo,
        'linkedin_pct': linkedin_pct,
        'cv_pct': cv_pct,
        'video_pct': video_pct,
        'profile_completo_count': profile_completo_count,
        'profile_completo_pct': profile_completo_pct,
        'sin_perfil_completo': sin_perfil_completo,
        'badge_data': badge_data,
        'badge_total': badge_total,
        'badge_tipos': badge_tipos,
        'alumnos_sin_badge': alumnos_sin_badge,
        'carreras_dist': carreras_dist,
        'region_dist': region_dist,
        'enps_empresa_avg': enps_empresa_avg,
        'user_rol': user_rol,
    }
    return render(request, 'pages/dashboard.html', context)


def proyectos_view(request):
    user_rol = _get_user_rol(request)
    # Aislamiento de período respetando el Modo Auditoría del administrador.
    periodo, periodo_label = _periodo_contextual(request)

    real_rol = getattr(request.user, 'rol', None)
    es_tutor_empresa = real_rol == 'tutor_empresa'
    empresa_tutor = None
    if es_tutor_empresa:
        perfil_tutor = TutorEmpresa.objects.select_related('empresa').filter(pk=request.user.pk).first()
        empresa_tutor = perfil_tutor.empresa if perfil_tutor else None

    # Vinculación contextual: al llegar desde el detalle de una empresa, el campo
    # de empresa viene fijo y oculto (se inyecta su id en el formulario).
    empresa_fija = None
    empresa_fija_id = (request.GET.get('empresa_fija') or '').strip()
    if empresa_fija_id and not es_tutor_empresa:
        empresa_fija = Empresa.objects.filter(id=empresa_fija_id).first()

    if request.method == 'POST' and request.POST.get('action') == 'create_proyecto':
        # Solo el administrador y el tutor empresa pueden crear proyectos.
        if real_rol not in ('admin', 'tutor_empresa'):
            return HttpResponseForbidden('No tienes permisos para crear proyectos.')
        titulo = (request.POST.get('titulo') or '').strip()[:255]
        if es_tutor_empresa:
            # El tutor de empresa no elige empresa: se asigna forzosamente la
            # empresa vinculada a su perfil, ignorando cualquier empresa_id enviado.
            empresa = empresa_tutor
        else:
            empresa_id = request.POST.get('empresa_id')
            empresa = Empresa.objects.filter(id=empresa_id).first() if empresa_id else None
        if not titulo or empresa is None:
            messages.error(request, 'El título y la empresa son obligatorios.')
        else:
            modalidad = (request.POST.get('modalidad') or '').strip().lower() or None
            if modalidad not in (None, 'presencial', 'hibrido', 'remoto'):
                modalidad = None
            carrera = Carrera.objects.filter(id=request.POST.get('carrera_id')).first() if request.POST.get('carrera_id') else None
            try:
                vacantes = int(request.POST.get('vacantes') or 1)
            except ValueError:
                vacantes = 1
            try:
                proyecto_nuevo = Proyecto.objects.create(
                    empresa=empresa,
                    titulo=titulo,
                    descripcion=(request.POST.get('descripcion') or '').strip() or None,
                    carrera=carrera,
                    modalidad=modalidad,
                    vacantes=max(1, vacantes),
                    is_active=True,
                    created_by=request.user,
                    created_at=timezone.now(),
                    updated_at=timezone.now(),
                )
                messages.success(request, f'Proyecto “{titulo}” creado correctamente.')
                # El admin aterriza directamente en el Hub del proyecto recién creado.
                if not es_tutor_empresa:
                    return redirect('proyecto_detalle', proyecto_id=proyecto_nuevo.id)
            except Exception as exc:
                messages.error(request, f'No se pudo crear el proyecto: {exc}')
        return redirect('proyectos')

    q = (request.GET.get('q') or '').strip()
    estado_filter = (request.GET.get('estado') or '').strip().lower()
    modalidad_filter = (request.GET.get('modalidad') or '').strip().lower()
    rubro_filter = (request.GET.get('rubro') or '').strip().lower()
    empresa_filter = (request.GET.get('empresa') or '').strip().lower()
    carrera_filter = (request.GET.get('carrera') or '').strip().lower()
    sort_filter = (request.GET.get('sort') or 'reciente').strip().lower()

    base_qs = (
        Proyecto.objects.select_related('empresa', 'carrera')
        .annotate(
            # Conteos aislados al período contextual: no se mezclan registros de
            # otros semestres en la vista del tutor ni del administrador.
            postulaciones_count=Count(
                'postulacion',
                filter=Q(postulacion__periodo=periodo),
                distinct=True,
            ),
            asignados_count=Count(
                'proyectoperiodo',
                filter=Q(proyectoperiodo__alumno__isnull=False, proyectoperiodo__periodo=periodo),
                distinct=True,
            ),
        )
        .order_by('-created_at', 'titulo')
    )

    modalidad_options = _build_filter_options(base_qs.values_list('modalidad', flat=True).distinct())
    rubro_options = _build_filter_options(base_qs.values_list('empresa__rubro', flat=True).distinct())
    empresa_options = _build_filter_options(base_qs.values_list('empresa__nombre', flat=True).distinct())
    carrera_options = _build_filter_options(base_qs.values_list('carrera__nombre', flat=True).distinct())
    proyectos = base_qs

    if q:
        proyectos = proyectos.filter(
            Q(titulo__icontains=q) | Q(empresa__nombre__icontains=q) | Q(carrera__nombre__icontains=q)
        )
    if estado_filter == 'abierto':
        proyectos = proyectos.filter(is_active=True)
    elif estado_filter == 'cerrado':
        proyectos = proyectos.filter(is_active=False)
    if modalidad_filter:
        proyectos = proyectos.filter(modalidad__icontains=modalidad_filter)
    if rubro_filter:
        proyectos = proyectos.filter(empresa__rubro__icontains=rubro_filter)
    if empresa_filter:
        proyectos = proyectos.filter(empresa__nombre__icontains=empresa_filter)
    if carrera_filter:
        proyectos = proyectos.filter(carrera__nombre__icontains=carrera_filter)

    if sort_filter == 'empresa':
        proyectos = proyectos.order_by('empresa__nombre', 'titulo')
    elif sort_filter == 'postulaciones':
        proyectos = proyectos.order_by('-postulaciones_count', 'titulo')
    elif sort_filter == 'vacantes':
        proyectos = proyectos.order_by('-vacantes', 'titulo')

    rows = []
    for p in proyectos:
        modalidad_label = (p.modalidad or 'Sin modalidad').strip().title()
        rubro_label = p.empresa.rubro or 'Sin rubro'
        vacantes = p.vacantes or 0
        disponibles = max(vacantes - p.asignados_count, 0)
        empresa_nombre = p.empresa.nombre or ''
        empresa_initials = ''.join([w[0] for w in empresa_nombre.split()[:2]]).upper() or 'EM'
        rows.append(
            {
                'id': p.id,
                'titulo': p.titulo,
                'empresa': empresa_nombre,
                'empresa_initials': empresa_initials,
                'empresa_logo_style': _empresa_logo_style(empresa_nombre),
                'descripcion': (p.descripcion or '')[:200],
                'rubro': rubro_label,
                'rubro_class': _rubro_class(rubro_label),
                'carrera': p.carrera.nombre if p.carrera else 'Multicarrera',
                'modalidad': modalidad_label,
                'modalidad_class': 'presencial' if 'presencial' in modalidad_label.lower() else 'hibrido' if 'hibrid' in modalidad_label.lower() or 'híbrid' in modalidad_label.lower() else 'remoto' if 'remoto' in modalidad_label.lower() else '',
                'vacantes': vacantes,
                'asignados': p.asignados_count,
                'disponibles': disponibles,
                'postulaciones': p.postulaciones_count,
                'is_active': bool(p.is_active),
                'estado': 'Abierto' if p.is_active else 'Cerrado',
                # Valores crudos para prellenar el modal de edición.
                'empresa_id': p.empresa_id,
                'carrera_id': p.carrera_id or '',
                'modalidad_raw': p.modalidad or '',
                'descripcion_full': p.descripcion or '',
            }
        )

    paginator = Paginator(rows, 25)
    page_obj = paginator.get_page(request.GET.get('page', 1))
    context = {
        'periodo_label': periodo_label,
        'proyectos': page_obj.object_list,
        'page_obj': page_obj,
        'proyectos_count': len(rows),
        'proyectos_activos': sum(1 for x in rows if x['is_active']),
        'filters': {
            'q': q,
            'estado': estado_filter,
            'modalidad': modalidad_filter,
            'rubro': rubro_filter,
            'empresa': empresa_filter,
            'carrera': carrera_filter,
            'sort': sort_filter,
        },
        'filter_options': {
            'modalidad': modalidad_options,
            'rubro': rubro_options,
            'empresa': empresa_options,
            'carrera': carrera_options,
        },
        'empresas_select': list(Empresa.objects.order_by('nombre').values('id', 'nombre')),
        'carreras_select': list(Carrera.objects.order_by('nombre').values('id', 'nombre')),
        'es_tutor_empresa': es_tutor_empresa,
        'empresa_tutor_nombre': empresa_tutor.nombre if empresa_tutor else '',
        'empresa_fija_id': empresa_fija.id if empresa_fija else '',
        'empresa_fija_nombre': empresa_fija.nombre if empresa_fija else '',
        'abrir_modal': bool(empresa_fija),
        'user_rol': user_rol,
    }
    return render(request, 'pages/proyectos.html', context)


def _alumno_dict(alumno, pp_id=None):
    """Representación ligera de un alumno para el Hub del proyecto."""
    u = alumno.id
    nombre = f"{u.nombre} {u.apellido}".strip()
    return {
        'pp_id': pp_id,
        'id': str(alumno.pk),
        'nombre': nombre,
        'email': u.email or '',
        'iniciales': (''.join(p[0] for p in nombre.split()[:2]).upper() or 'AL'),
        'carrera': alumno.carrera.nombre if alumno.carrera else 'Sin carrera',
    }


@require_roles('tutor_empresa', 'tutor_udd')
def proyecto_detalle_view(request, proyecto_id):
    """Hub del proyecto: datos base, tutores y alumnos, con asignación dinámica sobre ProyectoPeriodo."""
    user_rol = _get_user_rol(request)
    # Respeta el Modo Auditoría: si el admin previsualiza un período histórico, las
    # asignaciones (ProyectoPeriodo) se crean y editan sobre ESE período.
    periodo, periodo_label = _periodo_contextual(request)
    proyecto = Proyecto.objects.select_related('empresa', 'carrera').filter(id=proyecto_id).first()
    if proyecto is None:
        messages.error(request, 'El proyecto solicitado no existe.')
        return redirect('proyectos')

    # Un tutor empresa solo puede gestionar proyectos de su propia empresa.
    if getattr(request.user, 'rol', None) == 'tutor_empresa':
        perfil_te = TutorEmpresa.objects.filter(pk=request.user.pk).first()
        if perfil_te is None or proyecto.empresa_id != perfil_te.empresa_id:
            messages.error(request, 'Solo puedes gestionar proyectos de tu empresa.')
            return redirect('proyectos')

    if periodo is None:
        return render(request, 'pages/proyecto_detalle.html', {
            'user_rol': user_rol,
            'periodo_label': periodo_label,
            'proyecto': proyecto,
            'sin_periodo': True,
        })

    destino = redirect('proyecto_detalle', proyecto_id=proyecto.id)

    if request.method == 'POST':
        action = (request.POST.get('action') or '').strip()

        if action == 'asignar_alumno':
            alumno = Alumno.objects.select_related('id').filter(pk=(request.POST.get('alumno_id') or '').strip()).first()
            if alumno is None or not alumno.id.is_active:
                messages.error(request, 'Selecciona un alumno activo válido.')
                return destino
            # El alumno debe pertenecer a la empresa del proyecto (AsignacionPeriodo).
            pertenece = AsignacionPeriodo.objects.filter(
                alumno=alumno, periodo=periodo, empresa=proyecto.empresa, estado='activo'
            ).exists()
            if not pertenece:
                messages.error(request, 'Ese alumno no pertenece a la empresa del proyecto.')
                return destino
            if ProyectoPeriodo.objects.filter(periodo=periodo, alumno_id=alumno.pk).exists():
                messages.error(request, 'Ese alumno ya tiene un proyecto asignado en el período.')
                return destino
            asignados_actuales = ProyectoPeriodo.objects.filter(
                proyecto=proyecto, periodo=periodo, alumno__isnull=False
            ).count()
            if asignados_actuales >= (proyecto.vacantes or 0):
                messages.error(request, 'El proyecto ya alcanzó su capacidad de vacantes.')
                return destino
            placeholder = ProyectoPeriodo.objects.filter(
                proyecto=proyecto, periodo=periodo, alumno__isnull=True
            ).first()
            ahora = timezone.now()
            if placeholder:
                placeholder.alumno = alumno
                placeholder.estado = placeholder.estado or 'en_curso'
                placeholder.semana_actual = placeholder.semana_actual or 1
                placeholder.updated_at = ahora
                placeholder.save()
            else:
                nueva = ProyectoPeriodo.objects.create(
                    proyecto=proyecto,
                    periodo=periodo,
                    alumno=alumno,
                    estado='en_curso',
                    semana_actual=1,
                    created_at=ahora,
                    updated_at=ahora,
                )
                # Hereda el equipo de tutores del proyecto (otra fila del período).
                equipo = ProyectoPeriodo.objects.filter(
                    proyecto=proyecto, periodo=periodo
                ).exclude(pk=nueva.pk).first()
                if equipo:
                    nueva.tutores_udd.set(equipo.tutores_udd.all())
                    nueva.tutores_empresa.set(equipo.tutores_empresa.all())
            messages.success(request, 'Alumno añadido al proyecto.')
            return destino

        if action == 'cambiar_proyecto':
            pp = ProyectoPeriodo.objects.filter(pk=(request.POST.get('pp_id') or '').strip(), proyecto=proyecto, periodo=periodo).first()
            destino_proy = Proyecto.objects.filter(pk=(request.POST.get('proyecto_destino_id') or '').strip(), is_active=True).first()
            if pp is None or not pp.alumno_id or destino_proy is None:
                messages.error(request, 'Selecciona un proyecto de destino válido.')
                return destino
            if destino_proy.pk == proyecto.pk:
                messages.error(request, 'El alumno ya pertenece a este proyecto.')
                return destino
            asignados_destino = ProyectoPeriodo.objects.filter(
                proyecto=destino_proy, periodo=periodo, alumno__isnull=False
            ).count()
            if asignados_destino >= (destino_proy.vacantes or 0):
                messages.error(request, 'El proyecto de destino no tiene cupos disponibles.')
                return destino
            # Al trasladar al alumno, adopta el equipo de tutores del proyecto destino.
            equipo_destino = ProyectoPeriodo.objects.filter(
                proyecto=destino_proy, periodo=periodo
            ).exclude(pk=pp.pk).first()
            pp.proyecto = destino_proy
            pp.updated_at = timezone.now()
            pp.save()
            if equipo_destino:
                pp.tutores_udd.set(equipo_destino.tutores_udd.all())
                pp.tutores_empresa.set(equipo_destino.tutores_empresa.all())
            else:
                pp.tutores_udd.clear()
                pp.tutores_empresa.clear()
            messages.success(request, 'Alumno trasladado a “%s”.' % destino_proy.titulo)
            return destino

        if action == 'desvincular_alumno':
            pp = ProyectoPeriodo.objects.filter(pk=(request.POST.get('pp_id') or '').strip(), proyecto=proyecto, periodo=periodo).first()
            if pp is None or not pp.alumno_id:
                messages.error(request, 'No se encontró la asignación a desvincular.')
                return destino
            # Se libera al alumno conservando la fila (equipo e histórico intactos).
            pp.alumno = None
            pp.updated_at = timezone.now()
            pp.save()
            messages.success(request, 'Alumno desvinculado del proyecto.')
            return destino

        if action in ('asignar_tutor', 'desvincular_tutor'):
            tipo = (request.POST.get('tipo') or '').strip().lower()
            if tipo not in ('udd', 'empresa'):
                messages.error(request, 'Tipo de tutor no válido.')
                return destino

            etiqueta = 'UDD' if tipo == 'udd' else 'Empresa'
            tutor_id = (request.POST.get('tutor_id') or '').strip()
            if tipo == 'udd':
                tutor = TutorUdd.objects.filter(pk=tutor_id).first()
            else:
                tutor = TutorEmpresa.objects.filter(pk=tutor_id, empresa=proyecto.empresa).first()
            if tutor is None:
                messages.error(request, 'Selecciona un tutor válido.')
                return destino

            filas = list(ProyectoPeriodo.objects.filter(proyecto=proyecto, periodo=periodo))
            ahora = timezone.now()

            # Equipo actual del tipo en el proyecto-período. Todas las filas comparten
            # el mismo equipo, pero se calcula la unión por robustez.
            asignados = set()
            for fila in filas:
                relacion = fila.tutores_udd if tipo == 'udd' else fila.tutores_empresa
                asignados.update(relacion.values_list('pk', flat=True))

            if action == 'asignar_tutor':
                # ── Control de capacidad: máximo 2 tutores por tipo y proyecto ──
                if tutor.pk in asignados:
                    messages.error(request, 'Ese tutor %s ya está asignado a este proyecto.' % etiqueta)
                    return destino
                if len(asignados) >= 2:
                    messages.error(
                        request,
                        'No se puede asignar: se alcanzó el límite máximo de 2 tutores %s para este proyecto.' % etiqueta,
                    )
                    return destino
                if not filas:
                    # Sin filas aún: se crea una fila contenedora (sin alumno) del equipo.
                    filas = [ProyectoPeriodo.objects.create(
                        proyecto=proyecto, periodo=periodo,
                        estado='en_curso', created_at=ahora, updated_at=ahora,
                    )]
                for fila in filas:
                    relacion = fila.tutores_udd if tipo == 'udd' else fila.tutores_empresa
                    relacion.add(tutor)
                    fila.updated_at = ahora
                    fila.save()
                messages.success(request, 'Tutor %s asignado al proyecto.' % etiqueta)
            else:
                # desvincular_tutor: quita ese tutor de todas las filas del proyecto.
                for fila in filas:
                    relacion = fila.tutores_udd if tipo == 'udd' else fila.tutores_empresa
                    relacion.remove(tutor)
                    fila.updated_at = ahora
                    fila.save()
                messages.success(request, 'Tutor %s desvinculado del proyecto.' % etiqueta)
            return destino

        messages.error(request, 'Acción no reconocida.')
        return destino

    # ── GET: armado del Hub ──
    filas = list(
        ProyectoPeriodo.objects.filter(proyecto=proyecto, periodo=periodo)
        .select_related('alumno__id', 'alumno__carrera')
        .prefetch_related('tutores_udd__id', 'tutores_empresa__id')
    )
    alumnos = [_alumno_dict(f.alumno, pp_id=f.pk) for f in filas if f.alumno_id]
    alumnos.sort(key=lambda a: a['nombre'])

    # Equipo de tutores del proyecto (hasta 2 por tipo). Las filas comparten equipo;
    # se deduplica por si alguna quedó desincronizada.
    tutores_udd_asignados = []
    tutores_empresa_asignados = []
    udd_vistos = set()
    emp_vistos = set()
    for f in filas:
        for t in f.tutores_udd.all():
            if t.pk in udd_vistos:
                continue
            udd_vistos.add(t.pk)
            u = t.id
            nombre = f"{u.nombre} {u.apellido}".strip()
            tutores_udd_asignados.append({
                'id': str(t.pk),
                'nombre': nombre,
                'email': u.email or '',
                'iniciales': (''.join(p[0] for p in nombre.split()[:2]).upper() or 'TU'),
                'detalle': t.departamento or 'Tutor UDD',
            })
        for t in f.tutores_empresa.all():
            if t.pk in emp_vistos:
                continue
            emp_vistos.add(t.pk)
            u = t.id
            nombre = f"{u.nombre} {u.apellido}".strip()
            tutores_empresa_asignados.append({
                'id': str(t.pk),
                'nombre': nombre,
                'email': u.email or '',
                'iniciales': (''.join(p[0] for p in nombre.split()[:2]).upper() or 'TE'),
                'detalle': t.cargo or 'Tutor empresa',
            })

    # Alumnos activos sin proyecto en el período (para añadir / reasignar).
    asignados_ids = set(
        ProyectoPeriodo.objects.filter(periodo=periodo, alumno__isnull=False).values_list('alumno_id', flat=True)
    )
    # Solo alumnos cuya empresa (en AsignacionPeriodo del período) es la del proyecto.
    disponibles_qs = (
        Alumno.objects.select_related('id', 'carrera')
        .filter(
            id__is_active=True,
            asignaciones_periodo__periodo=periodo,
            asignaciones_periodo__empresa=proyecto.empresa,
            asignaciones_periodo__estado='activo',
        )
        .exclude(pk__in=asignados_ids)
        .distinct()
        .order_by('id__nombre', 'id__apellido')
    )
    disponibles = [
        {'id': str(a.pk), 'nombre': f"{a.id.nombre} {a.id.apellido}".strip(), 'email': a.id.email or '', 'carrera': a.carrera.nombre if a.carrera else 'Sin carrera'}
        for a in disponibles_qs
    ]

    # Proyectos de destino para "Cambiar Proyecto": activos, con cupos libres.
    asignados_por_proyecto = {
        row['proyecto']: row['c']
        for row in ProyectoPeriodo.objects.filter(periodo=periodo, alumno__isnull=False).values('proyecto').annotate(c=Count('id'))
    }
    proyectos_destino = []
    for p in Proyecto.objects.filter(is_active=True).exclude(id=proyecto.id).select_related('empresa').order_by('titulo'):
        libres = (p.vacantes or 0) - asignados_por_proyecto.get(p.id, 0)
        if libres > 0:
            proyectos_destino.append({
                'id': p.id,
                'titulo': p.titulo,
                'empresa': p.empresa.nombre if p.empresa else 'Sin empresa',
                'disponibles': libres,
            })

    tutores_udd = [
        {'id': str(t.pk), 'nombre': f"{t.id.nombre} {t.id.apellido}".strip(), 'detalle': t.departamento or 'Tutor UDD'}
        for t in TutorUdd.objects.select_related('id').order_by('id__nombre', 'id__apellido')
    ]
    tutores_empresa = [
        {'id': str(t.pk), 'nombre': f"{t.id.nombre} {t.id.apellido}".strip(), 'detalle': t.cargo or 'Tutor empresa'}
        for t in TutorEmpresa.objects.select_related('id').filter(empresa=proyecto.empresa).order_by('id__nombre', 'id__apellido')
    ]

    modalidad_label = (proyecto.modalidad or 'Sin definir').strip().title()
    context = {
        'user_rol': user_rol,
        'periodo_label': periodo_label,
        'sin_periodo': False,
        'proyecto': proyecto,
        'empresa_nombre': proyecto.empresa.nombre if proyecto.empresa else 'Sin empresa',
        'carrera_nombre': proyecto.carrera.nombre if proyecto.carrera else 'Multicarrera',
        'modalidad_label': modalidad_label,
        'descripcion': proyecto.descripcion or '',
        'vacantes': proyecto.vacantes or 0,
        'is_active': bool(proyecto.is_active),
        'alumnos': alumnos,
        'alumnos_count': len(alumnos),
        'cupos_llenos': len(alumnos) >= (proyecto.vacantes or 0),
        'tutores_udd_asignados': tutores_udd_asignados,
        'tutores_empresa_asignados': tutores_empresa_asignados,
        'tutor_udd_lleno': len(tutores_udd_asignados) >= 2,
        'tutor_empresa_lleno': len(tutores_empresa_asignados) >= 2,
        'disponibles': disponibles,
        'proyectos_destino': proyectos_destino,
        'tutores_udd': tutores_udd,
        'tutores_empresa': tutores_empresa,
    }
    return render(request, 'pages/proyecto_detalle.html', context)


def empresas_view(request):
    user_rol = _get_user_rol(request)
    # Aislamiento de período respetando el Modo Auditoría del administrador.
    periodo, periodo_label = _periodo_contextual(request)

    if request.method == 'POST' and request.POST.get('action') == 'create_empresa':
        # Solo el administrador puede crear empresas.
        if getattr(request.user, 'rol', None) != 'admin':
            return HttpResponseForbidden('No tienes permisos para crear empresas.')
        nombre = (request.POST.get('nombre') or '').strip()[:255]
        if not nombre:
            messages.error(request, 'El nombre de la empresa es obligatorio.')
        else:
            presencia = (request.POST.get('presencia') or '').strip().lower() or None
            if presencia not in (None, 'chile', 'multinacional'):
                presencia = None
            try:
                Empresa.objects.create(
                    nombre=nombre,
                    rubro=(request.POST.get('rubro') or '').strip()[:150] or None,
                    ubicacion=(request.POST.get('ubicacion') or '').strip()[:255] or None,
                    presencia=presencia,
                    descripcion=(request.POST.get('descripcion') or '').strip() or None,
                    contacto_nombre=(request.POST.get('contacto_nombre') or '').strip()[:150] or None,
                    contacto_email=(request.POST.get('contacto_email') or '').strip()[:254] or None,
                    is_active=True,
                )
                messages.success(request, f'Empresa “{nombre}” creada correctamente.')
            except Exception as exc:
                messages.error(request, f'No se pudo crear la empresa: {exc}')
        return redirect('empresas')

    q = (request.GET.get('q') or '').strip()
    rubro_filter = (request.GET.get('rubro') or '').strip().lower()
    presencia_filter = (request.GET.get('presencia') or '').strip().lower()
    sort_filter = (request.GET.get('sort') or 'practicantes').strip().lower()

    # Proyectos de la empresa con asignación en el período contextual, pre-buscados
    # con Prefetch para que el panel de detalle itere esta relación ya filtrada y no
    # dispare consultas extra ni muestre proyectos de otros semestres.
    proyectos_periodo_qs = (
        Proyecto.objects.filter(
            proyectoperiodo__periodo=periodo,
            proyectoperiodo__alumno__isnull=False,
        )
        .distinct()
        .order_by('-is_active', 'titulo')
    )
    base_qs = (
        Empresa.objects.annotate(
            proyectos_count=Count(
                'proyecto',
                filter=Q(proyecto__proyectoperiodo__periodo=periodo),
                distinct=True,
            ),
            practicantes_count=Count(
                'proyecto__proyectoperiodo',
                filter=Q(
                    proyecto__proyectoperiodo__alumno__isnull=False,
                    proyecto__proyectoperiodo__periodo=periodo,
                ),
                distinct=True,
            ),
            tutores_count=Count('tutorempresa', distinct=True),
        )
        .prefetch_related(
            Prefetch('proyecto_set', queryset=proyectos_periodo_qs, to_attr='proyectos_periodo')
        )
        .order_by('-practicantes_count', 'nombre')
    )

    rubro_options = _build_filter_options(base_qs.values_list('rubro', flat=True).distinct())
    presencia_options = _build_filter_options(base_qs.values_list('presencia', flat=True).distinct())
    empresas_qs = base_qs

    if q:
        empresas_qs = empresas_qs.filter(
            Q(nombre__icontains=q) | Q(rubro__icontains=q) | Q(ubicacion__icontains=q)
        )
    if rubro_filter:
        empresas_qs = empresas_qs.filter(rubro__icontains=rubro_filter)
    if presencia_filter:
        empresas_qs = empresas_qs.filter(presencia__icontains=presencia_filter)

    if sort_filter == 'nombre':
        empresas_qs = empresas_qs.order_by('nombre')
    elif sort_filter == 'enps':
        empresas_qs = empresas_qs.order_by('-enps_score', 'nombre')
    elif sort_filter == 'proyectos':
        empresas_qs = empresas_qs.order_by('-proyectos_count', 'nombre')

    # ── eNPS real: se calcula desde EvaluacionEmpresa filtrando por el período
    # activo. Escala 1–10: promotores (>=9), detractores (<=6); pasivos (7–8).
    # eNPS = (% promotores − % detractores), en el rango −100 a 100.
    enps_map = {}
    if periodo:
        conteo = {}
        for ev in EvaluacionEmpresa.objects.filter(periodo=periodo).values('empresa', 'puntuacion'):
            bucket = conteo.setdefault(ev['empresa'], {'total': 0, 'promotores': 0, 'detractores': 0})
            bucket['total'] += 1
            if ev['puntuacion'] >= 9:
                bucket['promotores'] += 1
            elif ev['puntuacion'] <= 6:
                bucket['detractores'] += 1
        for emp_id, b in conteo.items():
            enps_map[emp_id] = round((b['promotores'] - b['detractores']) / b['total'] * 100) if b['total'] else 0

    empresas = []
    for e in empresas_qs:
        rubro = e.rubro or 'Sin rubro'
        nombre = e.nombre or ''
        initials = ''.join([p[0] for p in nombre.split()[:2]]).upper() or 'EM'
        empresas.append(
            {
                'id': e.id,
                'nombre': e.nombre,
                'initials': initials,
                'logo_style': _empresa_logo_style(nombre),
                'rubro': rubro,
                'rubro_raw': e.rubro or '',
                'rubro_class': _rubro_class(rubro),
                'presencia': (e.presencia or 'Sin definir').title(),
                'presencia_raw': e.presencia or '',
                # None cuando no hay evaluaciones, para distinguir en la plantilla
                # "sin datos" de un eNPS real de 0.
                'enps': enps_map.get(e.id),
                'proyectos': e.proyectos_count,
                'practicantes': e.practicantes_count,
                'tutores': e.tutores_count,
                # Itera la relación ya pre-filtrada por período (Prefetch).
                'proyectos_list': [
                    {'id': pr.id, 'titulo': pr.titulo, 'is_active': bool(pr.is_active)}
                    for pr in e.proyectos_periodo
                ],
                'descripcion': e.descripcion or '',
                'contacto_nombre': e.contacto_nombre or '',
                'contacto_email': e.contacto_email or '',
                'ubicacion': e.ubicacion or '',
                'empleados': e.empleados_aprox or '',
            }
        )

    # Promedio de eNPS solo sobre empresas con evaluaciones reales en el período.
    # Sin evaluaciones -> None, para mostrar un estado neutro ("—") en vez de 0.
    enps_evaluadas = [enps_map[k] for k in enps_map]
    enps_avg = round(sum(enps_evaluadas) / len(enps_evaluadas), 1) if enps_evaluadas else None

    paginator = Paginator(empresas, 25)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    # Tutores empresa asociados (vía FK) para el panel de detalle de la página actual.
    page_ids = [e['id'] for e in page_obj.object_list]
    tutores_map = {}
    for te in TutorEmpresa.objects.filter(empresa_id__in=page_ids).select_related('id'):
        u = te.id
        nombre_tutor = f"{u.nombre} {u.apellido}".strip()
        tutores_map.setdefault(te.empresa_id, []).append({
            'nombre': nombre_tutor,
            'initials': (''.join(p[0] for p in nombre_tutor.split()[:2]).upper() or 'TE'),
            'cargo': te.cargo or '',
            'email': u.email or '',
        })
    # Los proyectos del período ya vienen pre-buscados (Prefetch) en cada empresa;
    # aquí solo resta enlazar los tutores (relación por FK, no dependiente de período).
    for e in page_obj.object_list:
        e['tutores_list'] = tutores_map.get(e['id'], [])

    context = {
        'periodo_label': periodo_label,
        'empresas': page_obj.object_list,
        'page_obj': page_obj,
        'empresas_count': len(empresas),
        'enps_avg': enps_avg,
        'filters': {
            'q': q,
            'rubro': rubro_filter,
            'presencia': presencia_filter,
            'sort': sort_filter,
        },
        'filter_options': {
            'rubro': rubro_options,
            'presencia': presencia_options,
        },
        'user_rol': user_rol,
    }
    return render(request, 'pages/empresas.html', context)


def empresa_editar_view(request):
    """Actualiza los datos de una empresa existente (solo administrador)."""
    if request.method != 'POST':
        return redirect('empresas')
    if getattr(request.user, 'rol', None) != 'admin':
        return HttpResponseForbidden('No tienes permisos para editar empresas.')

    empresa = Empresa.objects.filter(pk=(request.POST.get('empresa_id') or '').strip()).first()
    if empresa is None:
        messages.error(request, 'La empresa que intentas editar no existe.')
        return redirect('empresas')

    nombre = (request.POST.get('nombre') or '').strip()[:255]
    if not nombre:
        messages.error(request, 'El nombre de la empresa es obligatorio.')
        return redirect('empresas')

    presencia = (request.POST.get('presencia') or '').strip().lower() or None
    if presencia not in (None, 'chile', 'multinacional'):
        presencia = None

    try:
        empresa.nombre = nombre
        empresa.rubro = (request.POST.get('rubro') or '').strip()[:150] or None
        empresa.ubicacion = (request.POST.get('ubicacion') or '').strip()[:255] or None
        empresa.presencia = presencia
        empresa.descripcion = (request.POST.get('descripcion') or '').strip() or None
        empresa.contacto_nombre = (request.POST.get('contacto_nombre') or '').strip()[:150] or None
        empresa.contacto_email = (request.POST.get('contacto_email') or '').strip()[:254] or None
        empresa.updated_at = timezone.now()
        empresa.save()
        messages.success(request, f'Empresa “{nombre}” actualizada correctamente.')
    except Exception as exc:
        messages.error(request, f'No se pudo actualizar la empresa: {exc}')
    return redirect('empresas')


def empresa_eliminar_view(request):
    """Elimina una empresa solo si no tiene proyectos, practicantes ni tutores."""
    if request.method != 'POST':
        return redirect('empresas')
    if getattr(request.user, 'rol', None) != 'admin':
        return HttpResponseForbidden('No tienes permisos para eliminar empresas.')

    empresa = Empresa.objects.filter(pk=(request.POST.get('empresa_id') or '').strip()).first()
    if empresa is None:
        messages.error(request, 'La empresa que intentas eliminar no existe.')
        return redirect('empresas')

    nombre = empresa.nombre or 'la empresa'
    proyectos_count = Proyecto.objects.filter(empresa=empresa).count()
    practicantes_count = ProyectoPeriodo.objects.filter(
        proyecto__empresa=empresa, alumno__isnull=False
    ).count()
    tutores_count = TutorEmpresa.objects.filter(empresa=empresa).count()

    if proyectos_count or practicantes_count or tutores_count:
        messages.error(
            request,
            'No se puede eliminar “%s”: tiene %d proyecto(s), %d practicante(s) y %d tutor(es) '
            'asociados. Desvincula o reasigna esos registros antes de eliminarla.'
            % (nombre, proyectos_count, practicantes_count, tutores_count),
        )
        return redirect('empresas')

    try:
        empresa.delete()
        messages.success(request, f'Empresa “{nombre}” eliminada correctamente.')
    except IntegrityError:
        messages.error(
            request,
            'No se pudo eliminar “%s” porque conserva información asociada en otras tablas.' % nombre,
        )
    return redirect('empresas')


def proyecto_editar_view(request):
    """Actualiza un proyecto existente (solo administrador)."""
    if request.method != 'POST':
        return redirect('proyectos')
    if getattr(request.user, 'rol', None) != 'admin':
        return HttpResponseForbidden('No tienes permisos para editar proyectos.')

    proyecto = Proyecto.objects.filter(pk=(request.POST.get('proyecto_id') or '').strip()).first()
    if proyecto is None:
        messages.error(request, 'El proyecto que intentas editar no existe.')
        return redirect('proyectos')

    titulo = (request.POST.get('titulo') or '').strip()[:255]
    if not titulo:
        messages.error(request, 'El título del proyecto es obligatorio.')
        return redirect('proyectos')

    empresa = Empresa.objects.filter(pk=(request.POST.get('empresa_id') or '').strip()).first()
    if empresa is None:
        messages.error(request, 'Selecciona una empresa válida para el proyecto.')
        return redirect('proyectos')

    carrera = Carrera.objects.filter(pk=(request.POST.get('carrera_id') or '').strip()).first()
    modalidad = (request.POST.get('modalidad') or '').strip().lower() or None
    if modalidad not in (None, 'presencial', 'hibrido', 'remoto'):
        modalidad = None
    try:
        vacantes = int(request.POST.get('vacantes') or 0)
    except (TypeError, ValueError):
        vacantes = 0
    vacantes = max(0, min(vacantes, 99))

    try:
        proyecto.titulo = titulo
        proyecto.empresa = empresa
        proyecto.carrera = carrera
        proyecto.modalidad = modalidad
        proyecto.vacantes = vacantes
        proyecto.descripcion = (request.POST.get('descripcion') or '').strip() or None
        proyecto.is_active = request.POST.get('is_active') == 'on'
        proyecto.updated_at = timezone.now()
        proyecto.save()
        messages.success(request, f'Proyecto “{titulo}” actualizado correctamente.')
    except Exception as exc:
        messages.error(request, f'No se pudo actualizar el proyecto: {exc}')
    return redirect('proyectos')


def proyecto_eliminar_view(request):
    """Elimina un proyecto solo si no tiene postulaciones ni asignaciones de período."""
    if request.method != 'POST':
        return redirect('proyectos')
    if getattr(request.user, 'rol', None) != 'admin':
        return HttpResponseForbidden('No tienes permisos para eliminar proyectos.')

    proyecto = Proyecto.objects.filter(pk=(request.POST.get('proyecto_id') or '').strip()).first()
    if proyecto is None:
        messages.error(request, 'El proyecto que intentas eliminar no existe.')
        return redirect('proyectos')

    titulo = proyecto.titulo or 'el proyecto'
    postulaciones_count = Postulacion.objects.filter(proyecto=proyecto).count()
    periodos_count = ProyectoPeriodo.objects.filter(proyecto=proyecto).count()

    if postulaciones_count or periodos_count:
        messages.error(
            request,
            'No se puede eliminar “%s”: tiene %d postulación(es) y %d asignación(es) de período '
            'asociadas. Cierra o reasigna esos registros antes de eliminarlo.'
            % (titulo, postulaciones_count, periodos_count),
        )
        return redirect('proyectos')

    try:
        proyecto.delete()
        messages.success(request, f'Proyecto “{titulo}” eliminado correctamente.')
    except IntegrityError:
        messages.error(
            request,
            'No se pudo eliminar “%s” porque conserva información asociada en otras tablas.' % titulo,
        )
    return redirect('proyectos')


def alumno_editar_view(request):
    """Actualiza el perfil de un alumno y su usuario asociado (solo administrador)."""
    if request.method != 'POST':
        return redirect('alumnos')
    if getattr(request.user, 'rol', None) != 'admin':
        return HttpResponseForbidden('No tienes permisos para editar alumnos.')

    alumno = Alumno.objects.select_related('id').filter(pk=(request.POST.get('alumno_id') or '').strip()).first()
    if alumno is None:
        messages.error(request, 'El alumno que intentas editar no existe.')
        return redirect('alumnos')

    nombre = (request.POST.get('nombre') or '').strip()[:150]
    apellido = (request.POST.get('apellido') or '').strip()[:150]
    email = (request.POST.get('email') or '').strip()[:254]
    if not nombre or not email:
        messages.error(request, 'El nombre y el email del alumno son obligatorios.')
        return redirect('alumnos')
    if Usuario.objects.filter(email=email).exclude(pk=alumno.id_id).exists():
        messages.error(request, 'Ya existe otro usuario registrado con ese email.')
        return redirect('alumnos')

    carrera = Carrera.objects.filter(pk=(request.POST.get('carrera_id') or '').strip()).first()
    sede = Sede.objects.filter(pk=(request.POST.get('sede_id') or '').strip()).first()
    generacion = (request.POST.get('generacion') or '').strip()
    try:
        generacion = int(generacion) if generacion else None
    except ValueError:
        generacion = None

    nombre_full = f"{nombre} {apellido}".strip()
    try:
        usuario = alumno.id
        usuario.nombre = nombre
        usuario.apellido = apellido
        usuario.email = email
        usuario.updated_at = timezone.now()
        usuario.save()

        alumno.carrera = carrera
        alumno.sede = sede
        alumno.generacion = generacion
        alumno.matricula = (request.POST.get('matricula') or '').strip()[:30] or None
        # Enlaces de empleabilidad (alimentan la métrica del dashboard).
        alumno.url_linkedin = (request.POST.get('url_linkedin') or '').strip() or None
        alumno.url_cv = (request.POST.get('url_cv') or '').strip() or None
        alumno.url_youtube = (request.POST.get('url_youtube') or '').strip() or None
        alumno.updated_at = timezone.now()
        alumno.save()
        messages.success(request, f'Alumno “{nombre_full}” actualizado correctamente.')
    except Exception as exc:
        messages.error(request, f'No se pudo actualizar el alumno: {exc}')
    return redirect('alumnos')


def alumno_eliminar_view(request):
    """Elimina el perfil de un alumno solo si no tiene registros académicos; conserva el Usuario."""
    if request.method != 'POST':
        return redirect('alumnos')
    if getattr(request.user, 'rol', None) != 'admin':
        return HttpResponseForbidden('No tienes permisos para eliminar alumnos.')

    alumno = Alumno.objects.select_related('id').filter(pk=(request.POST.get('alumno_id') or '').strip()).first()
    if alumno is None:
        messages.error(request, 'El alumno que intentas eliminar no existe.')
        return redirect('alumnos')

    nombre = f"{alumno.id.nombre} {alumno.id.apellido}".strip() or 'el alumno'
    bitacoras_count = Bitacora.objects.filter(proyecto_periodo__alumno=alumno).count()
    postulaciones_count = Postulacion.objects.filter(alumno=alumno).count()
    periodos_count = ProyectoPeriodo.objects.filter(alumno=alumno).count()
    dependientes = (
        bitacoras_count
        + postulaciones_count
        + periodos_count
        + AlumnoBadge.objects.filter(alumno=alumno).count()
        + CalificacionCompetencia.objects.filter(alumno=alumno).count()
        + EvaluacionEmpresa.objects.filter(alumno=alumno).count()
    )

    if dependientes:
        messages.error(
            request,
            'No se puede eliminar a “%s”: tiene %d bitácora(s), %d postulación(es) y %d '
            'asignación(es) de período, además de otros registros académicos. Depura esos '
            'datos antes de eliminar al alumno.'
            % (nombre, bitacoras_count, postulaciones_count, periodos_count),
        )
        return redirect('alumnos')

    try:
        alumno.delete()
        messages.success(request, f'Alumno “{nombre}” eliminado correctamente.')
    except IntegrityError:
        messages.error(
            request,
            'No se pudo eliminar a “%s” porque conserva información asociada en otras tablas.' % nombre,
        )
    return redirect('alumnos')


def postulaciones_view(request):
    user_rol = _get_user_rol(request)
    # Aislamiento de período respetando el Modo Auditoría del administrador.
    periodo, periodo_label = _periodo_contextual(request)

    q = (request.GET.get('q') or '').strip()
    modalidad_filter = (request.GET.get('modalidad') or '').strip().lower()
    estado_filter = (request.GET.get('estado') or '').strip().lower()
    mis_estado_filter = (request.GET.get('mis_estado') or '').strip().lower()

    base_vacantes = (
        Proyecto.objects.select_related('empresa', 'carrera')
        .annotate(postulaciones_count=Count(
            'postulacion',
            filter=Q(postulacion__periodo=periodo),
            distinct=True,
        ))
        .order_by('-is_active', '-created_at', 'titulo')
    )

    modalidad_options = _build_filter_options(base_vacantes.values_list('modalidad', flat=True).distinct())
    vacantes = base_vacantes

    if q:
        vacantes = vacantes.filter(
            Q(titulo__icontains=q) | Q(empresa__nombre__icontains=q) | Q(carrera__nombre__icontains=q)
        )
    if modalidad_filter:
        vacantes = vacantes.filter(modalidad__icontains=modalidad_filter)
    if estado_filter == 'abierto':
        vacantes = vacantes.filter(is_active=True)
    elif estado_filter == 'cerrado':
        vacantes = vacantes.filter(is_active=False)

    vacantes_rows = []
    for p in vacantes:
        modalidad = (p.modalidad or 'Sin modalidad').strip().title()
        empresa_nombre = p.empresa.nombre or ''
        empresa_initials = ''.join([w[0] for w in empresa_nombre.split()[:2]]).upper() or 'EM'
        vacantes_rows.append(
            {
                'id': p.id,
                'titulo': p.titulo,
                'empresa': empresa_nombre,
                'empresa_initials': empresa_initials,
                'logo_style': _empresa_logo_style(empresa_nombre),
                'rubro': p.empresa.rubro or 'Sin rubro',
                'rubro_class': _rubro_class(p.empresa.rubro),
                'carrera': p.carrera.nombre if p.carrera else 'Multicarrera',
                'modalidad': modalidad,
                'vacantes': p.vacantes or 0,
                'postulaciones': p.postulaciones_count,
                'is_active': bool(p.is_active),
            }
        )

    alumno_ref = Alumno.objects.order_by('pk').first()
    mis_postulaciones = []
    if alumno_ref:
        mis_qs = (
            Postulacion.objects.select_related('proyecto__empresa', 'periodo')
            .filter(alumno=alumno_ref, periodo=periodo)
            .order_by('-fecha_postulacion')
        )
        for p in mis_qs:
            estado_item = (p.estado or 'en_revision').strip()
            empresa_nombre = p.proyecto.empresa.nombre or ''
            empresa_initials = ''.join([w[0] for w in empresa_nombre.split()[:2]]).upper() or 'EM'
            mis_postulaciones.append(
                {
                    'proyecto': p.proyecto.titulo,
                    'empresa': empresa_nombre,
                    'empresa_initials': empresa_initials,
                    'logo_style': _empresa_logo_style(empresa_nombre),
                    'fecha': p.fecha_postulacion,
                    'estado': estado_item,
                }
            )

    mis_estado_options = _build_filter_options([x['estado'] for x in mis_postulaciones])
    if mis_estado_filter:
        mis_postulaciones = [x for x in mis_postulaciones if (x['estado'] or '').lower() == mis_estado_filter]

    vacantes_abiertas = sum(1 for v in vacantes_rows if v['is_active'])

    pag_vac = Paginator(vacantes_rows, 25)
    page_obj_vacantes = pag_vac.get_page(request.GET.get('pv', 1))

    context = {
        'periodo_label': periodo_label,
        'vacantes': page_obj_vacantes.object_list,
        'page_obj_vacantes': page_obj_vacantes,
        'vacantes_count': len(vacantes_rows),
        'vacantes_abiertas': vacantes_abiertas,
        'mis_postulaciones': mis_postulaciones,
        'mis_postulaciones_count': len(mis_postulaciones),
        'filters': {
            'q': q,
            'modalidad': modalidad_filter,
            'estado': estado_filter,
            'mis_estado': mis_estado_filter,
        },
        'filter_options': {
            'modalidad': modalidad_options,
            'mis_estado': mis_estado_options,
        },
        'initial_tab': request.GET.get('tab', 'vacantes'),
        'user_rol': user_rol,
    }
    return render(request, 'pages/postulaciones.html', context)


@require_roles('alumno')
def postular_view(request):
    if request.method != 'POST':
        return redirect('postulaciones')
    proyecto_id = request.POST.get('proyecto_id')
    periodo, _ = _periodo_contextual(request)
    alumno_ref = Alumno.objects.order_by('pk').first()
    if proyecto_id and periodo and alumno_ref:
        try:
            proyecto = Proyecto.objects.get(id=proyecto_id, is_active=True)
            Postulacion.objects.get_or_create(
                proyecto=proyecto,
                alumno=alumno_ref,
                periodo=periodo,
                defaults={
                    'estado': 'en_revision',
                    'fecha_postulacion': timezone.now(),
                    'fecha_actualizacion': timezone.now(),
                },
            )
        except Exception:
            pass
    return redirect('/postulaciones/?tab=mis')


@require_roles()
def alumnos_view(request):
    user_rol = _get_user_rol(request)
    # Aislamiento de período respetando el Modo Auditoría del administrador.
    periodo, periodo_label = _periodo_contextual(request)

    q = (request.GET.get('q') or '').strip()
    carrera_filter = (request.GET.get('carrera') or '').strip().lower()
    sede_filter = (request.GET.get('sede') or '').strip().lower()
    perfil_min_filter = request.GET.get('perfil_min')

    from django.db.models import Max as _Max
    # La participación del alumno (proyecto, empresa y bitácoras) se cuenta solo
    # dentro del período contextual; los badges son logros históricos y no se acotan.
    base_qs = (
        Alumno.objects.select_related('id', 'carrera', 'sede')
        .annotate(
            badges_count=Count('alumnobadge', distinct=True),
            proyectos_count=Count(
                'proyectoperiodo',
                filter=Q(proyectoperiodo__periodo=periodo),
                distinct=True,
            ),
            bitacoras_count=Count(
                'proyectoperiodo__bitacora',
                filter=Q(proyectoperiodo__periodo=periodo),
                distinct=True,
            ),
            # La empresa proviene de AsignacionPeriodo del período contextual, de modo
            # que cambia correctamente al alternar el período en el Modo Auditoría.
            empresa_nombre=_Max(
                'asignaciones_periodo__empresa__nombre',
                filter=Q(asignaciones_periodo__periodo=periodo, asignaciones_periodo__estado='activo'),
            ),
        )
        .order_by('id__nombre', 'id__apellido')
    )

    carrera_options = _build_filter_options(base_qs.values_list('carrera__nombre', flat=True).distinct())
    sede_options = _build_filter_options(base_qs.values_list('sede__nombre', flat=True).distinct())
    alumnos_qs = base_qs

    if q:
        alumnos_qs = alumnos_qs.filter(
            Q(id__nombre__icontains=q)
            | Q(id__apellido__icontains=q)
            | Q(id__email__icontains=q)
            | Q(carrera__nombre__icontains=q)
        )
    if carrera_filter:
        alumnos_qs = alumnos_qs.filter(carrera__nombre__icontains=carrera_filter)
    if sede_filter:
        alumnos_qs = alumnos_qs.filter(sede__nombre__icontains=sede_filter)

    # ── Filtro por estado de asignación en el período contextual ──
    estado_filter = (request.GET.get('estado') or '').strip().lower()
    if estado_filter not in ('asignados', 'sin_asignar'):
        estado_filter = ''

    # Conteo de alumnos sin asignación (respeta la búsqueda y filtros vigentes) para
    # el indicador de la pestaña Sin Asignar.
    if periodo is not None:
        sin_asignar_count = alumnos_qs.exclude(proyectoperiodo__periodo=periodo).count()
    else:
        sin_asignar_count = alumnos_qs.count()

    if estado_filter == 'sin_asignar':
        # exclude: alumnos SIN ningún registro intermedio en el período activo.
        alumnos_qs = alumnos_qs.exclude(proyectoperiodo__periodo=periodo) if periodo is not None else alumnos_qs
    elif estado_filter == 'asignados':
        # filter: alumnos CON al menos un registro intermedio en el período activo.
        alumnos_qs = alumnos_qs.filter(proyectoperiodo__periodo=periodo).distinct() if periodo is not None else alumnos_qs.none()

    alumnos = []
    for a in alumnos_qs:
        linkedin_ok = bool(a.url_linkedin)
        cv_ok = bool(a.url_cv)
        video_ok = bool(a.url_youtube)
        profile_pct = int(((int(linkedin_ok) + int(cv_ok) + int(video_ok)) / 3) * 100)
        full_name = f"{a.id.nombre} {a.id.apellido}".strip()
        initials = ''.join([p[0] for p in full_name.split()[:2]]).upper() or 'AL'
        alumnos.append(
            {
                'id': str(a.id_id),
                'nombre': full_name,
                'initials': initials,
                'logo_style': _empresa_logo_style(full_name),
                'email': a.id.email,
                'carrera': a.carrera.nombre if a.carrera else 'Sin carrera',
                'sede': a.sede.nombre if a.sede else 'Sin sede',
                'generacion': a.generacion or '-',
                'profile_pct': profile_pct,
                'linkedin_ok': linkedin_ok,
                'cv_ok': cv_ok,
                'video_ok': video_ok,
                'badges_count': a.badges_count,
                'proyectos_count': a.proyectos_count,
                'bitacoras_count': a.bitacoras_count,
                'empresa': a.empresa_nombre or '',
                # Valores crudos para prellenar el modal de edición.
                'nombre_pila': a.id.nombre or '',
                'apellido': a.id.apellido or '',
                'carrera_id': a.carrera_id or '',
                'sede_id': a.sede_id or '',
                'generacion_raw': a.generacion or '',
                'matricula': a.matricula or '',
                'url_linkedin': a.url_linkedin or '',
                'url_cv': a.url_cv or '',
                'url_youtube': a.url_youtube or '',
            }
        )

    if perfil_min_filter:
        try:
            minimo = int(perfil_min_filter)
            alumnos = [x for x in alumnos if x['profile_pct'] >= minimo]
        except ValueError:
            pass

    profile_avg = round(sum(item['profile_pct'] for item in alumnos) / len(alumnos)) if alumnos else 0
    perfil_min_options = [str(pct) for pct in sorted({item['profile_pct'] for item in alumnos})]

    paginator = Paginator(alumnos, 25)
    page_obj = paginator.get_page(request.GET.get('page', 1))
    context = {
        'periodo_label': periodo_label,
        'alumnos': page_obj.object_list,
        'page_obj': page_obj,
        'alumnos_count': len(alumnos),
        'profile_avg': profile_avg,
        'filters': {
            'q': q,
            'carrera': carrera_filter,
            'sede': sede_filter,
            'perfil_min': perfil_min_filter or '',
            'estado': estado_filter,
        },
        'sin_asignar_count': sin_asignar_count,
        'filter_options': {
            'carrera': carrera_options,
            'sede': sede_options,
            'perfil_min': perfil_min_options,
        },
        'carreras_select': list(Carrera.objects.order_by('nombre').values('id', 'nombre')),
        'sedes_select': list(Sede.objects.order_by('nombre').values('id', 'nombre')),
        'user_rol': user_rol,
    }
    return render(request, 'pages/alumnos.html', context)


def _semana_estado_combinado(estado_emp, estado_udd, has_sent):
    """Estado consolidado de la semana para la línea de tiempo (timeline)."""
    ce = _bitacora_estado_class(estado_emp)
    cu = _bitacora_estado_class(estado_udd)
    if ce == 'aprobado' and cu == 'aprobado':
        return 'aprobada'
    if 'corregido' in (ce, cu):
        return 'corregida'
    if 'reprobado' in (ce, cu):
        return 'reprobada'
    if has_sent:
        return 'enviada'
    return 'pendiente'


@require_roles('tutor_udd', 'tutor_empresa')
def bitacoras_view(request):
    user_rol = _get_user_rol(request)
    # Aislamiento de período respetando el Modo Auditoría del administrador.
    periodo, periodo_label = _periodo_contextual(request)

    if not periodo:
        return render(request, 'pages/bitacoras.html', {
            'periodo_label': periodo_label,
            'user_rol': user_rol,
            'bitacoras_empty': True,
            'message': 'No hay un período activo para mostrar el seguimiento de bitácoras.',
        })

    project_periods = (
        ProyectoPeriodo.objects.filter(periodo=periodo, alumno__isnull=False)
        .select_related('proyecto__empresa', 'alumno__id')
        .prefetch_related('tutores_empresa__id', 'tutores_udd__id')
        .order_by('alumno__id__nombre', 'alumno__id__apellido')
    )

    alumno_uid = (request.GET.get('alumno') or '').strip()
    current_pp = None
    if alumno_uid:
        current_pp = project_periods.filter(alumno__id_id=alumno_uid).first()
    if current_pp is None:
        current_pp = project_periods.first()

    student_options = [
        {
            'id': str(pp.alumno.id_id),
            'label': f"{pp.alumno.id.nombre} {pp.alumno.id.apellido}".strip(),
            'empresa': pp.proyecto.empresa.nombre if pp.proyecto and pp.proyecto.empresa else '',
        }
        for pp in project_periods
    ]

    if current_pp is None:
        return render(request, 'pages/bitacoras.html', {
            'periodo_label': periodo_label,
            'user_rol': user_rol,
            'bitacoras_empty': True,
            'message': 'No hay alumnos con proyecto asignado en el período activo.',
        })

    total_semanas = periodo.total_semanas or 13
    bitacoras_by_week = {
        b.semana: b
        for b in Bitacora.objects.filter(proyecto_periodo=current_pp)
    }

    weeks = []
    semanas_aprobadas = 0
    for n in range(1, total_semanas + 1):
        b = bitacoras_by_week.get(n)
        estado_emp = b.estado_emp if b else None
        estado_udd = b.estado_udd if b else None
        has_sent = bool(b and b.fecha_envio)
        estado = _semana_estado_combinado(estado_emp, estado_udd, has_sent)
        if estado == 'aprobada':
            semanas_aprobadas += 1
        weeks.append({
            'n': n,
            'estado': estado,
            'estado_emp': _bitacora_estado_class(estado_emp),
            'estado_udd': _bitacora_estado_class(estado_udd),
            'estado_emp_label': _bitacora_estado_label(estado_emp),
            'estado_udd_label': _bitacora_estado_label(estado_udd),
            'texto': (b.texto or '') if b else '',
            'has_sent': has_sent,
            'readonly': estado in ('aprobada', 'enviada', 'corregida', 'reprobada'),
            'fecha_str': b.fecha_envio.strftime('%d/%m/%Y %H:%M') if (b and b.fecha_envio) else '',
            'feedback_emp': (b.feedback_emp or '') if b else '',
            'feedback_udd': (b.feedback_udd or '') if b else '',
        })

    # Nota final de bitácoras (escala chilena) según la exigencia configurada.
    config_eval = ConfigEvaluacionPeriodo.objects.filter(periodo=periodo).first()
    porcentaje_exigencia = config_eval.umbral_bitacoras_pct if (config_eval and config_eval.umbral_bitacoras_pct) else 60
    nota_final = calcular_nota_bitacoras(semanas_aprobadas, total_semanas, porcentaje_exigencia)
    cumplimiento_pct = round((semanas_aprobadas / total_semanas) * 100) if total_semanas else 0

    def _initials(nombre):
        return ''.join(p[0] for p in nombre.split()[:2]).upper() or '--'

    emp_users = [t.id for t in current_pp.tutores_empresa.all()]
    udd_users = [t.id for t in current_pp.tutores_udd.all()]
    tutor_empresa_nombre = ', '.join(f"{u.nombre} {u.apellido}".strip() for u in emp_users) or 'Sin asignar'
    tutor_udd_nombre = ', '.join(f"{u.nombre} {u.apellido}".strip() for u in udd_users) or 'Sin asignar'

    context = {
        'periodo_label': periodo_label,
        'bitacoras_empty': False,
        'student_options': student_options,
        'selected_student_id': str(current_pp.alumno.id_id),
        'student_name': f"{current_pp.alumno.id.nombre} {current_pp.alumno.id.apellido}".strip(),
        'student_company': current_pp.proyecto.empresa.nombre if current_pp.proyecto and current_pp.proyecto.empresa else 'Sin empresa',
        'project_name': current_pp.proyecto.titulo if current_pp.proyecto else 'Sin proyecto',
        'weeks': weeks,
        'total_semanas': total_semanas,
        'semanas_aprobadas': semanas_aprobadas,
        'cumplimiento_pct': cumplimiento_pct,
        'porcentaje_exigencia': porcentaje_exigencia,
        'nota_final': nota_final,
        'tutor_empresa_nombre': tutor_empresa_nombre,
        'tutor_empresa_initials': _initials(tutor_empresa_nombre),
        'tutor_udd_nombre': tutor_udd_nombre,
        'tutor_udd_initials': _initials(tutor_udd_nombre),
        'user_rol': user_rol,
    }
    return render(request, 'pages/bitacoras.html', context)


@require_roles('tutor_udd', 'tutor_empresa')
def evaluaciones_view(request):
    """Gestor de Hitos del período con métricas por hito (rúbricas contestadas y promedio)."""
    user_rol = _get_user_rol(request)
    # Aislamiento de período respetando el Modo Auditoría.
    periodo, periodo_label = _periodo_contextual(request)

    # ── Administrador: buscador de alumnos → detalle (boletín) por alumno ──
    # Sin alumno seleccionado se muestra el panel de búsqueda; al elegir uno se
    # renderiza el detalle de su evaluación (boletín de notas del período).
    if user_rol == 'admin' and periodo:
        alumno_uid = (request.GET.get('alumno') or '').strip()
        if not alumno_uid:
            return _render_buscador_alumno(
                request, periodo, periodo_label, user_rol,
                'Evaluaciones · Selección de alumno', 'Ver evaluación',
            )
        current_pp = (
            ProyectoPeriodo.objects.filter(
                periodo=periodo, alumno__isnull=False, alumno__id_id=alumno_uid
            )
            .select_related('proyecto__empresa', 'alumno__id')
            .first()
        )
        if current_pp is None:
            return _render_buscador_alumno(
                request, periodo, periodo_label, user_rol,
                'Evaluaciones · Selección de alumno', 'Ver evaluación',
            )
        rows, promedio_final, en_calculo, pendientes = _calcular_boletin(periodo, current_pp)
        return render(request, 'pages/mis_notas.html', {
            'user_rol': user_rol,
            'periodo_label': periodo_label,
            'sin_datos': False,
            'student_name': f"{current_pp.alumno.id.nombre} {current_pp.alumno.id.apellido}".strip(),
            'student_company': current_pp.proyecto.empresa.nombre if current_pp.proyecto and current_pp.proyecto.empresa else 'Sin empresa',
            'project_name': current_pp.proyecto.titulo if current_pp.proyecto else 'Sin proyecto',
            'rows': rows,
            'promedio_final': promedio_final,
            'en_calculo': en_calculo,
            'pendientes': pendientes,
            'volver_url': request.path,
        })

    hitos_resumen = []
    total_rubricas = 0
    suma_global = 0.0
    cuenta_global = 0

    if periodo:
        total_alumnos = ProyectoPeriodo.objects.filter(periodo=periodo, alumno__isnull=False).count()

        # Agregados por hito desde las evaluaciones (rúbricas) contestadas.
        agregados = {
            row['hito']: row
            for row in (
                EvaluacionHito.objects.filter(hito__periodo=periodo)
                .values('hito')
                .annotate(cantidad=Count('id'), promedio=Avg('nota_calculada'))
            )
        }

        eval_labels = {'udd': 'Tutor UDD', 'empresa': 'Tutor Empresa', 'ambos': 'Ambos tutores'}
        for h in HitoEvaluacion.objects.filter(periodo=periodo).order_by('orden', 'semana'):
            datos = agregados.get(h.pk)
            cantidad = datos['cantidad'] if datos else 0
            promedio = round(float(datos['promedio']), 1) if (datos and datos['promedio'] is not None) else None
            total_rubricas += cantidad
            if promedio is not None:
                suma_global += float(datos['promedio']) * cantidad
                cuenta_global += cantidad
            evaluador = (h.evaluador or 'ambos').strip().lower()
            hitos_resumen.append({
                'id': h.pk,
                'nombre': h.nombre,
                'semana': h.semana,
                'peso_pct': h.peso_pct,
                'evaluador_label': eval_labels.get(evaluador, 'Ambos tutores'),
                'cantidad': cantidad,
                'total_alumnos': total_alumnos,
                'pct': int((cantidad / total_alumnos) * 100) if total_alumnos else 0,
                'promedio': promedio,
            })

    promedio_global = round(suma_global / cuenta_global, 1) if cuenta_global else None

    context = {
        'periodo_label': periodo_label,
        'hitos_resumen': hitos_resumen,
        'hitos_count': len(hitos_resumen),
        'total_rubricas': total_rubricas,
        'promedio_global': promedio_global,
        'user_rol': user_rol,
    }
    return render(request, 'pages/evaluaciones.html', context)


@require_roles()
def hitos_config_view(request):
    user_rol = _get_user_rol(request)
    # Respeta el Modo Auditoría: los hitos y la configuración de evaluación se crean
    # y editan sobre el período histórico que el admin esté previsualizando.
    periodo, periodo_label = _periodo_contextual(request)
    saved_ok = False

    if request.method == 'POST' and periodo:
        action = request.POST.get('action', '')
        if action == 'save_periodo':
            tipo_ciclo = request.POST.get('tipo_ciclo', 'semestral').strip()
            if tipo_ciclo in ('semestral', 'trimestral'):
                periodo.tipo_ciclo = tipo_ciclo
                try:
                    periodo.save(update_fields=['tipo_ciclo'])
                    saved_ok = True
                except Exception:
                    pass
        elif action == 'save_bitacoras':
            config, _ = ConfigEvaluacionPeriodo.objects.get_or_create(
                periodo=periodo,
                defaults={'peso_bitacoras_pct': 0, 'created_at': timezone.now()},
            )
            try:
                peso = int(request.POST.get('peso_bitacoras', config.peso_bitacoras_pct) or 0)
                config.peso_bitacoras_pct = max(0, min(100, peso))
                config.metodo_calculo_bitacoras = request.POST.get('metodo_bitacoras', config.metodo_calculo_bitacoras or 'aprobadas')
                umbral = int(request.POST.get('umbral_bitacoras', config.umbral_bitacoras_pct or 80) or 80)
                config.umbral_bitacoras_pct = max(0, min(100, umbral))
                config.updated_at = timezone.now()
                config.save()
                saved_ok = True
            except Exception:
                pass
        elif action == 'save_fechas':
            from datetime import datetime as _dt
            fi_str = request.POST.get('fecha_inicio', '').strip()
            ff_str = request.POST.get('fecha_fin', '').strip()
            try:
                fi = _dt.strptime(fi_str, '%Y-%m-%d').date()
                ff = _dt.strptime(ff_str, '%Y-%m-%d').date()
                if fi <= ff:
                    periodo.fecha_inicio = fi
                    periodo.fecha_fin = ff
                    periodo.total_semanas = max(1, round((ff - fi).days / 7))
                    periodo.save(update_fields=['fecha_inicio', 'fecha_fin', 'total_semanas'])
                    saved_ok = True
            except Exception:
                pass
        elif action == 'delete_hito':
            hito_id = request.POST.get('hito_id')
            try:
                HitoEvaluacion.objects.filter(id=hito_id, periodo=periodo).delete()
                saved_ok = True
            except Exception:
                pass
        elif action == 'add_hito':
            nombre = request.POST.get('nombre', '').strip()
            try:
                semana = int(request.POST.get('semana') or 1)
                peso = int(request.POST.get('peso') or 0)
                evaluador = request.POST.get('evaluador', 'tutor_udd').strip()
                if nombre:
                    max_ord = HitoEvaluacion.objects.filter(periodo=periodo).aggregate(Max('orden'))['orden__max'] or 0
                    HitoEvaluacion.objects.create(
                        periodo=periodo,
                        nombre=nombre,
                        semana=semana,
                        peso_pct=max(0, min(100, peso)),
                        evaluador=evaluador,
                        estado='activo',
                        orden=max_ord + 1,
                        created_at=timezone.now(),
                        updated_at=timezone.now(),
                    )
                    saved_ok = True
            except Exception:
                pass
        elif action == 'edit_hito':
            hito_id = request.POST.get('hito_id')
            nombre = request.POST.get('nombre', '').strip()
            try:
                semana = int(request.POST.get('semana') or 1)
                peso = int(request.POST.get('peso') or 0)
                evaluador = request.POST.get('evaluador', 'tutor_udd').strip()
                h = HitoEvaluacion.objects.get(id=hito_id, periodo=periodo)
                if nombre:
                    h.nombre = nombre
                h.semana = semana
                h.peso_pct = max(0, min(100, peso))
                h.evaluador = evaluador
                h.updated_at = timezone.now()
                h.save(update_fields=['nombre', 'semana', 'peso_pct', 'evaluador', 'updated_at'])
                saved_ok = True
            except Exception:
                pass

    estado_filter = (request.GET.get('estado') or '').strip().lower()
    evaluador_filter = (request.GET.get('evaluador') or '').strip().lower()
    q = (request.GET.get('q') or '').strip()

    config = ConfigEvaluacionPeriodo.objects.filter(periodo=periodo).first() if periodo else None
    base_qs = (
        HitoEvaluacion.objects.filter(periodo=periodo).order_by('orden', 'semana')
        if periodo
        else HitoEvaluacion.objects.none()
    )
    estado_options = _build_filter_options(base_qs.values_list('estado', flat=True).distinct())
    evaluador_options = _build_filter_options(base_qs.values_list('evaluador', flat=True).distinct())
    hitos_qs = base_qs

    if estado_filter:
        hitos_qs = hitos_qs.filter(estado__icontains=estado_filter)
    if evaluador_filter:
        hitos_qs = hitos_qs.filter(evaluador__icontains=evaluador_filter)
    if q:
        hitos_qs = hitos_qs.filter(nombre__icontains=q)

    hitos = []
    for h in hitos_qs:
        competencias = h.competenciahito_set.order_by('orden', 'nombre')
        hitos.append(
            {
                'id': h.id,
                'nombre': h.nombre,
                'semana': h.semana,
                'peso': h.peso_pct,
                'evaluador': (h.evaluador or 'Sin definir').title(),
                'estado': (h.estado or 'activo').title(),
                'competencias': [
                    {'nombre': c.nombre, 'peso': c.peso_pct}
                    for c in competencias
                ],
            }
        )

    total_peso_hitos = sum(h['peso'] for h in hitos)
    peso_bitacoras = config.peso_bitacoras_pct if config else 0

    tipo_ciclo = getattr(periodo, 'tipo_ciclo', 'semestral') or 'semestral'
    total_peso = total_peso_hitos + peso_bitacoras
    avg_comps = round(sum(len(h['competencias']) for h in hitos) / len(hitos), 1) if hitos else 0

    context = {
        'periodo_label': periodo_label,
        'periodo': periodo,
        'tipo_ciclo': tipo_ciclo,
        'hitos': hitos,
        'hitos_count': len(hitos),
        'peso_hitos': total_peso_hitos,
        'peso_bitacoras': peso_bitacoras,
        'total_peso': total_peso,
        'avg_comps': avg_comps,
        'metodo_bitacoras': (config.metodo_calculo_bitacoras if config else 'aprobadas'),
        'umbral_bitacoras': (config.umbral_bitacoras_pct if config else 80),
        'saved_ok': saved_ok,
        'filters': {
            'estado': estado_filter,
            'evaluador': evaluador_filter,
            'q': q,
        },
        'filter_options': {
            'estado': estado_options,
            'evaluador': evaluador_options,
        },
        'user_rol': user_rol,
    }
    return render(request, 'pages/hitos_config.html', context)


# ─── CONFIGURACIÓN GLOBAL (períodos / evaluaciones / badges) ─────────────────

ICONOS_BADGE = ['🏆', '⭐', '🎯', '🚀', '💡', '🤝', '📈', '🔥', '🎖️', '👑']


@require_roles()
def configuracion_view(request):
    """Panel de configuración global con 3 áreas: períodos, evaluaciones, badges."""
    import json
    from datetime import datetime as _dt

    user_rol = _get_user_rol(request)
    active_tab = 'periodos'
    # Formulario de alta de período: queda enlazado solo si un POST de creación
    # falla la validación, para repintar los errores junto a los campos.
    periodo_form = None

    if request.method == 'POST':
        action = request.POST.get('action', '')
        active_tab = request.POST.get('active_tab', 'periodos')

        # ── Período académico: ALTA de un nuevo registro (nunca actualiza) ──
        if action == 'save_periodo':
            # Instancia en blanco (sin pasar el período activo): fuerza un INSERT.
            form = PeriodoAcademicoForm(request.POST)
            if form.is_valid():
                periodo = form.save(commit=False)
                # Las fechas ICLAE determinan el número de semanas de bitácora.
                periodo.total_semanas = max(1, round(
                    (periodo.fecha_fin_iclae - periodo.fecha_inicio_iclae).days / 7
                ))
                periodo.is_active = False
                periodo.created_at = timezone.now()
                periodo.save()
                messages.success(
                    request,
                    f'Período “{periodo.nombre}” creado con {periodo.total_semanas} semanas. '
                    'Actívalo desde la tabla cuando corresponda.'
                )
                # PRG: redirige tras el éxito para limpiar el formulario y evitar
                # el reenvío al recargar la página.
                return redirect(f"{request.path}?tab=periodos")
            # Validación fallida: conserva el formulario enlazado y cae al render
            # del GET para mostrar los errores sin perder lo que el usuario escribió.
            periodo_form = form
            active_tab = 'periodos'

        # ── Evaluación continua (bitácoras): porcentaje de exigencia fijo ──
        elif action == 'save_exigencia_bitacoras':
            periodo = PeriodoAcademico.objects.filter(is_active=True).first()
            if periodo is None:
                messages.error(request, 'Primero crea un período activo en la pestaña Períodos.')
            else:
                try:
                    exigencia = int(request.POST.get('porcentaje_exigencia') or 60)
                except ValueError:
                    exigencia = 60
                exigencia = max(1, min(99, exigencia))
                config, _ = ConfigEvaluacionPeriodo.objects.get_or_create(
                    periodo=periodo,
                    defaults={'peso_bitacoras_pct': 0, 'created_at': timezone.now()},
                )
                config.umbral_bitacoras_pct = exigencia
                config.updated_at = timezone.now()
                config.save()
                messages.success(request, f'Exigencia de bitácoras guardada en {exigencia}%.')
            return redirect(f"{request.path}?tab=evaluaciones")

        # ── Hito de evaluación con competencias dinámicas ──
        elif action == 'add_hito':
            periodo = PeriodoAcademico.objects.filter(is_active=True).first()
            nombre = (request.POST.get('nombre') or '').strip()[:150]
            try:
                semana = int(request.POST.get('semana') or 1)
            except ValueError:
                semana = 1
            evaluador = (request.POST.get('evaluador') or 'udd').strip()
            if evaluador not in ('udd', 'empresa', 'ambos'):
                evaluador = 'udd'
            try:
                competencias = json.loads(request.POST.get('competencias_json') or '[]')
            except json.JSONDecodeError:
                competencias = []

            comps = [
                {'nombre': str(c.get('nombre', '')).strip()[:150], 'peso': int(c.get('peso') or 0)}
                for c in competencias
                if str(c.get('nombre', '')).strip()
            ]
            suma = sum(c['peso'] for c in comps)

            if periodo is None:
                messages.error(request, 'Primero crea un período activo en la pestaña Períodos.')
            elif not nombre or not comps:
                messages.error(request, 'Ingresa el nombre del hito y al menos una competencia.')
            elif suma != 100:
                messages.error(request, f'La ponderación de las competencias debe sumar 100% (suma actual: {suma}%).')
            else:
                max_ord = HitoEvaluacion.objects.filter(periodo=periodo).aggregate(Max('orden'))['orden__max'] or 0
                hito = HitoEvaluacion.objects.create(
                    periodo=periodo,
                    nombre=nombre,
                    semana=semana,
                    peso_pct=100,
                    evaluador=evaluador,
                    estado='pendiente',
                    orden=max_ord + 1,
                    created_at=timezone.now(),
                    updated_at=timezone.now(),
                )
                for orden, c in enumerate(comps, start=1):
                    CompetenciaHito.objects.create(
                        hito=hito,
                        nombre=c['nombre'],
                        peso_pct=c['peso'],
                        orden=orden,
                        created_at=timezone.now(),
                    )
                messages.success(request, f'Hito “{nombre}” creado con {len(comps)} competencias.')
            return redirect(f"{request.path}?tab=evaluaciones")

        elif action == 'delete_hito':
            try:
                HitoEvaluacion.objects.filter(id=request.POST.get('hito_id')).delete()
                messages.success(request, 'Hito eliminado.')
            except Exception:
                messages.error(request, 'No se pudo eliminar el hito.')
            return redirect(f"{request.path}?tab=evaluaciones")

        # ── Badges / insignias ──
        elif action == 'add_badge':
            nombre = (request.POST.get('nombre') or '').strip()[:100]
            descripcion = (request.POST.get('descripcion') or '').strip()
            icono = (request.POST.get('icono') or '🏆').strip()[:10]
            tipo = (request.POST.get('tipo_otorgamiento') or 'manual').strip()
            if tipo not in ('automatico', 'manual'):
                tipo = 'manual'
            if not nombre:
                messages.error(request, 'El nombre de la insignia es obligatorio.')
            else:
                Badge.objects.create(
                    nombre=nombre,
                    descripcion=descripcion or None,
                    icono=icono,
                    criterio=tipo,
                    is_active=True,
                    created_at=timezone.now(),
                )
                messages.success(request, f'Insignia “{nombre}” creada.')
            return redirect(f"{request.path}?tab=badges")

        elif action == 'delete_badge':
            try:
                Badge.objects.filter(id=request.POST.get('badge_id')).delete()
                messages.success(request, 'Insignia eliminada.')
            except Exception:
                messages.error(request, 'No se pudo eliminar la insignia.')
            return redirect(f"{request.path}?tab=badges")

    # ── GET ──
    tab_param = (request.GET.get('tab') or active_tab).strip().lower()
    if tab_param not in ('institucion', 'periodos', 'evaluaciones', 'badges'):
        tab_param = 'periodos'

    periodo = PeriodoAcademico.objects.filter(is_active=True).first()
    periodo_label = periodo.nombre if periodo else 'Sin periodo activo'

    config_eval = ConfigEvaluacionPeriodo.objects.filter(periodo=periodo).first() if periodo else None
    porcentaje_exigencia = config_eval.umbral_bitacoras_pct if (config_eval and config_eval.umbral_bitacoras_pct) else 60

    eval_labels = {'udd': 'Tutor UDD', 'empresa': 'Tutor Empresa', 'ambos': 'Ambos tutores'}
    hitos = []
    if periodo:
        for h in HitoEvaluacion.objects.filter(periodo=periodo).order_by('orden', 'semana'):
            hitos.append({
                'id': h.id,
                'nombre': h.nombre,
                'semana': h.semana,
                'evaluador': (h.evaluador or 'udd'),
                'evaluador_label': eval_labels.get(h.evaluador or 'udd', 'Sin definir'),
                'competencias': [
                    {'nombre': c.nombre, 'peso': c.peso_pct}
                    for c in h.competenciahito_set.order_by('orden', 'nombre')
                ],
            })

    badges = [
        {
            'id': b.id,
            'nombre': b.nombre,
            'descripcion': b.descripcion or '',
            'icono': b.icono or '🏆',
            'tipo': (b.criterio or 'manual'),
            'tipo_label': 'Automático por hito' if (b.criterio or '') == 'automatico' else 'Manual por tutor',
        }
        for b in Badge.objects.order_by('-id')
    ]

    # ── Institución: sedes y carreras (formulario ultrasimplificado: solo nombre) ──
    sedes = [{'id': s.id, 'nombre': s.nombre} for s in Sede.objects.order_by('nombre')]
    carreras = [{'id': c.id, 'nombre': c.nombre} for c in Carrera.objects.order_by('nombre')]

    # ── Períodos académicos: listado completo para la tabla CRUD ──
    periodos = [
        {
            'id': p.id,
            'nombre': p.nombre,
            'fecha_inicio': p.fecha_inicio.strftime('%Y-%m-%d') if p.fecha_inicio else '',
            'fecha_fin': p.fecha_fin.strftime('%Y-%m-%d') if p.fecha_fin else '',
            'fecha_inicio_iclae': p.fecha_inicio_iclae.strftime('%Y-%m-%d') if p.fecha_inicio_iclae else '',
            'fecha_fin_iclae': p.fecha_fin_iclae.strftime('%Y-%m-%d') if p.fecha_fin_iclae else '',
            'total_semanas': p.total_semanas,
            'is_active': bool(p.is_active),
        }
        # Historial completo: todos los períodos, del más reciente al más antiguo.
        for p in PeriodoAcademico.objects.all().order_by('-fecha_inicio')
    ]

    context = {
        'periodo_label': periodo_label,
        'periodo': periodo,
        'active_tab': tab_param,
        'hitos': hitos,
        'hitos_count': len(hitos),
        'badges': badges,
        'badges_count': len(badges),
        'iconos_badge': ICONOS_BADGE,
        'porcentaje_exigencia': porcentaje_exigencia,
        'sedes': sedes,
        'sedes_count': len(sedes),
        'carreras': carreras,
        'carreras_count': len(carreras),
        'periodos': periodos,
        'periodos_count': len(periodos),
        'periodo_form': periodo_form or PeriodoAcademicoForm(),
        'user_rol': user_rol,
    }
    return render(request, 'pages/configuracion.html', context)


def _universidad_por_defecto():
    """Universidad para la FK obligatoria de Sede/Carrera: la primera, o una mínima si no hay."""
    universidad = Universidad.objects.order_by('pk').first()
    if universidad is None:
        universidad = Universidad.objects.create(nombre='Universidad del Desarrollo', codigo='UDD')
    return universidad


def sede_guardar_view(request):
    """Crea o actualiza una Sede (solo nombre) desde Configuración. Solo admin."""
    if request.method != 'POST':
        return redirect('configuracion')
    if getattr(request.user, 'rol', None) != 'admin':
        return HttpResponseForbidden('No tienes permisos para gestionar sedes.')

    destino = f"{reverse('configuracion')}?tab=institucion"
    nombre = (request.POST.get('nombre') or '').strip()[:150]
    if not nombre:
        messages.error(request, 'El nombre de la sede es obligatorio.')
        return redirect(destino)

    sede_id = (request.POST.get('sede_id') or '').strip()
    sede = Sede.objects.filter(pk=sede_id).first() if sede_id else Sede()
    if sede_id and sede is None:
        messages.error(request, 'La sede que intentas editar no existe.')
        return redirect(destino)

    try:
        sede.nombre = nombre
        # FK obligatoria por diseño previo: se completa por defecto si falta.
        if sede.universidad_id is None:
            sede.universidad = _universidad_por_defecto()
        sede.save()
        messages.success(request, f'Sede “{nombre}” guardada correctamente.')
    except Exception as exc:
        messages.error(request, f'No se pudo guardar la sede: {exc}')
    return redirect(destino)


def sede_eliminar_view(request):
    """Elimina una Sede solo si no tiene alumnos, tutores ni proyectos vinculados."""
    if request.method != 'POST':
        return redirect('configuracion')
    if getattr(request.user, 'rol', None) != 'admin':
        return HttpResponseForbidden('No tienes permisos para eliminar sedes.')

    destino = f"{reverse('configuracion')}?tab=institucion"
    sede = Sede.objects.filter(pk=(request.POST.get('sede_id') or '').strip()).first()
    if sede is None:
        messages.error(request, 'La sede que intentas eliminar no existe.')
        return redirect(destino)

    nombre = sede.nombre or 'la sede'
    alumnos_count = Alumno.objects.filter(sede=sede).count()
    tutores_count = TutorUdd.objects.filter(sede=sede).count()
    periodos_count = ProyectoPeriodo.objects.filter(sede=sede).count()

    if alumnos_count or tutores_count or periodos_count:
        messages.error(
            request,
            'No se puede eliminar “%s”: tiene %d alumno(s), %d tutor(es) UDD y %d '
            'asignación(es) de período vinculadas. Reasigna esos registros antes de eliminarla.'
            % (nombre, alumnos_count, tutores_count, periodos_count),
        )
        return redirect(destino)

    try:
        sede.delete()
        messages.success(request, f'Sede “{nombre}” eliminada correctamente.')
    except IntegrityError:
        messages.error(request, 'No se pudo eliminar “%s” porque conserva información asociada en otras tablas.' % nombre)
    return redirect(destino)


def carrera_guardar_view(request):
    """Crea o actualiza una Carrera (solo nombre) desde Configuración. Solo admin."""
    if request.method != 'POST':
        return redirect('configuracion')
    if getattr(request.user, 'rol', None) != 'admin':
        return HttpResponseForbidden('No tienes permisos para gestionar carreras.')

    destino = f"{reverse('configuracion')}?tab=institucion"
    nombre = (request.POST.get('nombre') or '').strip()[:255]
    if not nombre:
        messages.error(request, 'El nombre de la carrera es obligatorio.')
        return redirect(destino)

    carrera_id = (request.POST.get('carrera_id') or '').strip()
    carrera = Carrera.objects.filter(pk=carrera_id).first() if carrera_id else Carrera()
    if carrera_id and carrera is None:
        messages.error(request, 'La carrera que intentas editar no existe.')
        return redirect(destino)

    try:
        carrera.nombre = nombre
        # FK obligatoria por diseño previo: se completa por defecto si falta.
        if carrera.universidad_id is None:
            carrera.universidad = _universidad_por_defecto()
        carrera.save()
        messages.success(request, f'Carrera “{nombre}” guardada correctamente.')
    except Exception as exc:
        messages.error(request, f'No se pudo guardar la carrera: {exc}')
    return redirect(destino)


def carrera_eliminar_view(request):
    """Elimina una Carrera solo si no tiene alumnos ni proyectos vinculados."""
    if request.method != 'POST':
        return redirect('configuracion')
    if getattr(request.user, 'rol', None) != 'admin':
        return HttpResponseForbidden('No tienes permisos para eliminar carreras.')

    destino = f"{reverse('configuracion')}?tab=institucion"
    carrera = Carrera.objects.filter(pk=(request.POST.get('carrera_id') or '').strip()).first()
    if carrera is None:
        messages.error(request, 'La carrera que intentas eliminar no existe.')
        return redirect(destino)

    nombre = carrera.nombre or 'la carrera'
    alumnos_count = Alumno.objects.filter(carrera=carrera).count()
    proyectos_count = Proyecto.objects.filter(carrera=carrera).count()

    if alumnos_count or proyectos_count:
        messages.error(
            request,
            'No se puede eliminar “%s”: tiene %d alumno(s) y %d proyecto(s) vinculados. '
            'Reasigna esos registros antes de eliminarla.'
            % (nombre, alumnos_count, proyectos_count),
        )
        return redirect(destino)

    try:
        carrera.delete()
        messages.success(request, f'Carrera “{nombre}” eliminada correctamente.')
    except IntegrityError:
        messages.error(request, 'No se pudo eliminar “%s” porque conserva información asociada en otras tablas.' % nombre)
    return redirect(destino)


def periodo_editar_view(request):
    """Edita un PeriodoAcademico (nombre y fechas, recalcula total_semanas). Solo admin."""
    if request.method != 'POST':
        return redirect('configuracion')
    if getattr(request.user, 'rol', None) != 'admin':
        return HttpResponseForbidden('No tienes permisos para editar períodos.')

    from datetime import datetime as _dt

    destino = f"{reverse('configuracion')}?tab=periodos"
    periodo = PeriodoAcademico.objects.filter(pk=(request.POST.get('periodo_id') or '').strip()).first()
    if periodo is None:
        messages.error(request, 'El período que intentas editar no existe.')
        return redirect(destino)

    nombre = (request.POST.get('nombre') or '').strip()[:20]

    def _parse(name):
        raw = (request.POST.get(name) or '').strip()
        return _dt.strptime(raw, '%Y-%m-%d').date() if raw else None

    try:
        fi_acad = _parse('fecha_inicio')
        ff_acad = _parse('fecha_fin')
        fi_iclae = _parse('fecha_inicio_iclae')
        ff_iclae = _parse('fecha_fin_iclae')
    except ValueError:
        messages.error(request, 'Formato de fecha inválido.')
        return redirect(destino)

    if not nombre or not fi_acad or not ff_acad or not fi_iclae or not ff_iclae:
        messages.error(request, 'Completa el nombre y las cuatro fechas.')
        return redirect(destino)
    if fi_acad > ff_acad or fi_iclae > ff_iclae:
        messages.error(request, 'Las fechas de inicio no pueden ser posteriores a las de fin.')
        return redirect(destino)

    try:
        periodo.nombre = nombre
        periodo.fecha_inicio = fi_acad
        periodo.fecha_fin = ff_acad
        periodo.fecha_inicio_iclae = fi_iclae
        periodo.fecha_fin_iclae = ff_iclae
        periodo.total_semanas = max(1, round((ff_iclae - fi_iclae).days / 7))
        periodo.save()
        messages.success(request, f'Período “{nombre}” actualizado correctamente.')
    except Exception as exc:
        messages.error(request, f'No se pudo actualizar el período: {exc}')
    return redirect(destino)


def periodo_eliminar_view(request):
    """Elimina un PeriodoAcademico revalidando identidad (nombre, correo, contraseña) y deja auditoría."""
    if request.method != 'POST':
        return redirect('configuracion')
    if getattr(request.user, 'rol', None) != 'admin':
        return HttpResponseForbidden('No tienes permisos para eliminar períodos.')

    destino = f"{reverse('configuracion')}?tab=periodos"
    periodo = PeriodoAcademico.objects.filter(pk=(request.POST.get('periodo_id') or '').strip()).first()
    if periodo is None:
        messages.error(request, 'El período que intentas eliminar no existe.')
        return redirect(destino)

    nombre_in = (request.POST.get('nombre') or '').strip()
    correo_in = (request.POST.get('correo') or '').strip()
    password_in = request.POST.get('contrasena') or ''

    admin = request.user
    nombre_ok = bool(nombre_in) and nombre_in.casefold() == (admin.get_full_name() or '').strip().casefold()
    correo_ok = bool(correo_in) and correo_in.casefold() == (admin.email or '').strip().casefold()
    password_ok = bool(password_in) and admin.check_password(password_in)

    if not (nombre_ok and correo_ok and password_ok):
        messages.error(request, 'Verificación de seguridad fallida: nombre, correo o contraseña no coinciden. La eliminación fue rechazada.')
        return redirect(destino)

    periodo_nombre = periodo.nombre
    periodo_pk = periodo.pk
    try:
        periodo.delete()
    except IntegrityError:
        messages.error(
            request,
            'No se puede eliminar “%s” porque tiene información académica vinculada (evaluaciones, '
            'postulaciones o asignaciones de período). Depura esos registros antes de eliminarlo.'
            % periodo_nombre,
        )
        return redirect(destino)

    RegistroAuditoria.objects.create(
        administrador=admin,
        periodo_id=periodo_pk,
        periodo_nombre=periodo_nombre,
    )
    messages.success(request, f'Período “{periodo_nombre}” eliminado. La acción quedó registrada en auditoría.')
    return redirect(destino)


def periodo_activar_view(request):
    """Activa globalmente un período (desactiva el resto). Afecta a todos los usuarios."""
    if request.method != 'POST':
        return redirect('configuracion')
    if getattr(request.user, 'rol', None) != 'admin':
        return HttpResponseForbidden('No tienes permisos para activar períodos.')

    destino = f"{reverse('configuracion')}?tab=periodos"
    periodo = PeriodoAcademico.objects.filter(pk=(request.POST.get('periodo_id') or '').strip()).first()
    if periodo is None:
        messages.error(request, 'El período que intentas activar no existe.')
        return redirect(destino)

    PeriodoAcademico.objects.filter(is_active=True).exclude(pk=periodo.pk).update(is_active=False)
    periodo.is_active = True
    periodo.save(update_fields=['is_active'])
    messages.success(request, f'Período “{periodo.nombre}” activado globalmente. La vista de toda la plataforma se actualizó.')
    return redirect(destino)


def audit_periodo_view(request):
    """Modo Auditoría: guarda en sesión el período a previsualizar, sin tocar la BD."""
    if request.method != 'POST':
        return redirect(request.META.get('HTTP_REFERER') or 'dashboard')
    if getattr(request.user, 'rol', None) != 'admin':
        return HttpResponseForbidden('Solo un administrador puede usar el modo auditoría.')

    seleccion = (request.POST.get('audit_period_id') or '').strip()
    if not seleccion or seleccion == 'global':
        request.session.pop('audit_period_id', None)
    elif PeriodoAcademico.objects.filter(pk=seleccion).exists():
        request.session['audit_period_id'] = seleccion

    destino = request.POST.get('next') or request.META.get('HTTP_REFERER') or reverse('dashboard')
    return redirect(destino)


SEGMENTO_LABELS = {
    'todos': 'Todos los usuarios',
    'alumnos': 'Alumnos',
    'tutores_empresa': 'Tutores empresa',
    'tutores_udd': 'Tutores UDD',
}

# La tabla recordatorio_masivo es externa (managed=False) y no tiene columna de
# estado; el estado se infiere de cantidad_envios. Se usa este centinela negativo
# para marcar un envío Fallido sin alterar el esquema externo: None/0 = procesado,
# valor positivo = enviados reales, valor negativo = fallo del servidor de correo.
RECORDATORIO_FALLIDO = -1


def _dividir_anuncio(mensaje):
    """Separa el `mensaje` almacenado ("asunto: cuerpo") en (asunto, cuerpo)."""
    texto = mensaje or ''
    if ': ' in texto:
        asunto, cuerpo = texto.split(': ', 1)
        return asunto.strip(), cuerpo.strip()
    return texto.strip(), ''


@require_roles('alumno', 'tutor_udd', 'tutor_empresa')
def notificaciones_view(request):
    user_rol = _get_user_rol(request)
    # is_admin para el render usa el rol efectivo (respeta la previsualización
    # del admin); la protección del POST usa el rol REAL por seguridad.
    is_admin = user_rol == 'admin'
    real_rol = getattr(request.user, 'rol', None)
    _, periodo_label = _periodo_activo()
    sent_ok = False
    sent_count = 0

    if request.method == 'POST' and request.POST.get('action') == 'send_announcement':
        # Solo un administrador puede emitir nuevas notificaciones masivas.
        if real_rol != 'admin':
            return HttpResponseForbidden('No tienes permisos para enviar notificaciones.')
        segmento = (request.POST.get('segmento') or 'todos').strip()[:50]
        titulo = (request.POST.get('titulo') or '').strip()[:255]
        mensaje = (request.POST.get('mensaje') or '').strip()
        if titulo and mensaje:
            # 1) Se registra la campaña (cantidad_envios=0 = "Procesando").
            recordatorio = RecordatorioMasivo.objects.create(
                enviado_por=request.user,
                segmento=segmento,
                mensaje=f'{titulo}: {mensaje}',
                cantidad_envios=0,
                fecha_envio=timezone.now(),
            )
            # 2) Envío sincrónico: delegarlo a un worker inexistente dejaba la
            #    campaña en "Procesando" para siempre con 0 destinatarios. Se procesa
            #    aquí mismo y se captura cualquier fallo del servidor de correo.
            from .tasks import enviar_anuncio_masivo_task

            try:
                enviados = enviar_anuncio_masivo_task(recordatorio.pk, segmento, titulo, mensaje)
                if enviados:
                    messages.success(request, f'Comunicado enviado con éxito a {enviados} destinatario(s).')
                else:
                    messages.warning(
                        request,
                        'El comunicado se procesó, pero no hay destinatarios con participación '
                        'en el período en curso para ese segmento.'
                    )
                sent_ok = True
                sent_count = enviados
            except Exception as exc:
                # Se imprime en consola para depuración y la campaña queda en estado
                # "Fallido" en vez de quedar bloqueada en "Procesando".
                print(f'[Comunicados] Fallo al enviar el comunicado {recordatorio.pk}: {exc}')
                RecordatorioMasivo.objects.filter(pk=recordatorio.pk).update(
                    cantidad_envios=RECORDATORIO_FALLIDO
                )
                messages.error(request, 'Error al enviar el comunicado. Revisa la configuración del servidor.')
        else:
            messages.error(request, 'Completa el asunto y el mensaje del comunicado.')
        return redirect('notificaciones')

    if request.method == 'POST' and request.POST.get('action') == 'archivar_comunicado':
        # Borrado lógico: el comunicado se oculta de la bandeja del admin pero el
        # RecordatorioMasivo y las notificaciones entregadas permanecen intactos.
        if real_rol != 'admin':
            return HttpResponseForbidden('No tienes permisos para archivar comunicados.')
        try:
            recordatorio_id = int(request.POST.get('recordatorio_id') or 0)
        except (TypeError, ValueError):
            recordatorio_id = 0
        if recordatorio_id:
            RecordatorioArchivado.objects.get_or_create(
                recordatorio_id=recordatorio_id,
                defaults={'archivado_por': request.user},
            )
            messages.success(request, 'Comunicado archivado. Sigue visible para sus destinatarios.')
        return redirect('notificaciones')

    context = {
        'periodo_label': periodo_label,
        'sent_ok': sent_ok,
        'sent_count': sent_count,
        'user_rol': user_rol,
        'is_admin': is_admin,
    }

    if is_admin:
        # ── Admin: Comunicados Institucionales (campañas masivas emitidas) ──
        # Se excluyen los comunicados archivados (borrado lógico) de la vista.
        archivados = set(RecordatorioArchivado.objects.values_list('recordatorio_id', flat=True))
        recordatorios_qs = list(
            RecordatorioMasivo.objects.order_by('-fecha_envio')[:80]
        )
        recordatorios_visibles = [r for r in recordatorios_qs if r.pk not in archivados]
        campanas = []
        for r in recordatorios_visibles:
            asunto, _cuerpo = _dividir_anuncio(r.mensaje)
            cantidad = r.cantidad_envios
            if cantidad is not None and cantidad < 0:
                # Centinela negativo: el envío falló (ver RECORDATORIO_FALLIDO).
                estado_label, estado_class, cantidad_envios = 'Fallido', 'fallido', 0
            elif cantidad is None:
                estado_label, estado_class, cantidad_envios = 'Procesando', 'procesando', 0
            else:
                estado_label, estado_class, cantidad_envios = 'Enviado', 'enviado', cantidad
            campanas.append(
                {
                    'id': r.pk,
                    'fecha': r.fecha_envio,
                    'asunto': asunto or 'Anuncio sin asunto',
                    'segmento': SEGMENTO_LABELS.get(r.segmento, r.segmento or 'General'),
                    'cantidad_envios': cantidad_envios,
                    'estado_label': estado_label,
                    'estado_class': estado_class,
                }
            )
        context['campanas'] = campanas
        context['total_campanas'] = RecordatorioMasivo.objects.exclude(id__in=archivados).count()
        context['ultimo_envio'] = recordatorios_visibles[0].fecha_envio if recordatorios_visibles else None
    else:
        # ── Alumno / Tutor: Bandeja de Entrada personal ──
        # Se separan en dos listas: pendientes (no leídas) y leídas (histórico).
        notificaciones_qs = Notificacion.objects.filter(
            destinatario=request.user
        ).order_by('-fecha_envio', '-created_at')[:200]

        def _a_dict(n):
            return {
                'id': n.pk,
                'asunto': n.titulo or 'Mensaje sin asunto',
                'cuerpo': n.mensaje or '',
                'fecha': n.fecha_envio or n.created_at,
                'leida': bool(n.leida),
            }

        mensajes_pendientes = [_a_dict(n) for n in notificaciones_qs if not n.leida]
        mensajes_leidos = [_a_dict(n) for n in notificaciones_qs if n.leida]
        context['mensajes_pendientes'] = mensajes_pendientes
        context['mensajes_leidos'] = mensajes_leidos
        context['no_leidas'] = len(mensajes_pendientes)

    return render(request, 'pages/notificaciones.html', context)


@require_roles('alumno', 'tutor_udd', 'tutor_empresa')
def notificacion_leer_view(request, notif_id):
    """Marca como leída una notificación del usuario activo (fetch asíncrono)."""
    if request.method != 'POST':
        return HttpResponseForbidden('Método no permitido.')
    notificacion = Notificacion.objects.filter(pk=notif_id, destinatario=request.user).first()
    if notificacion is None:
        return JsonResponse({'ok': False, 'error': 'no encontrada'}, status=404)
    if not notificacion.leida:
        notificacion.leida = True
        notificacion.save(update_fields=['leida'])
    return JsonResponse({'ok': True})


@require_roles('alumno', 'tutor_udd', 'tutor_empresa')
def notificaciones_limpiar_view(request):
    """Limpia (oculta) las notificaciones ya leídas del usuario activo."""
    if request.method != 'POST':
        return redirect('notificaciones')
    eliminadas, _ = Notificacion.objects.filter(destinatario=request.user, leida=True).delete()
    if eliminadas:
        messages.success(request, 'Se limpiaron las notificaciones leídas.')
    else:
        messages.info(request, 'No tienes notificaciones leídas para limpiar.')
    return redirect('notificaciones')


# ─── EXCEL IMPORT ────────────────────────��────────────────────────��─────────

# Esquema de columnas por tipo de usuario importable. apellido no es obligatorio:
# si falta, se deriva del nombre completo con _separar_nombre_completo.
IMPORT_SCHEMAS = {
    'alumno': {
        'required': ['nombre', 'email'],
        'optional': ['apellido', 'carrera', 'sede', 'generacion', 'matricula', 'empresa'],
    },
    'tutor_udd': {
        'required': ['nombre', 'email'],
        'optional': ['apellido', 'sede', 'departamento'],
    },
    'tutor_empresa': {
        'required': ['nombre', 'email'],
        'optional': ['apellido', 'empresa', 'cargo'],
    },
}


# Conectores que se ignoran al comparar nombres de carrera (acentos aparte).
_CONECTORES_CARRERA = {'en', 'de', 'del', 'la', 'el', 'los', 'las', 'e', 'y'}


def _tokens_carrera(texto):
    """Palabras significativas de un nombre de carrera, sin acentos ni conectores."""
    return [t for t in _normalizar_texto(texto).split() if t not in _CONECTORES_CARRERA]


def _resolver_carrera(nombre):
    """Empareja la carrera tolerando acentos y conectores; la crea si no existe.

    Necesario porque la columna carrera_id del alumno es obligatoria en la base: si
    no se resolviera, el alumno no podría crearse. Primero busca por coincidencia de
    palabras significativas (así "Ingeniería Civil Informatica e Innovación
    Tecnologica" empareja con "...en Informática e Innovación Tecnológica") y, solo si
    no hay ninguna, crea la carrera con el nombre recibido.
    """
    nombre = (nombre or '').strip()
    if not nombre:
        return None
    objetivo = _tokens_carrera(nombre)
    if objetivo:
        for carrera in Carrera.objects.all():
            if _tokens_carrera(carrera.nombre) == objetivo:
                return carrera
    carrera, _ = Carrera.objects.get_or_create(
        nombre=nombre[:255],
        defaults={'universidad': _universidad_por_defecto()},
    )
    return carrera


def _resolver_sede(nombre):
    """Empareja la sede por nombre tolerando acentos; None si no se reconoce."""
    nombre = (nombre or '').strip()
    if not nombre:
        return None
    objetivo = _normalizar_texto(nombre)
    for sede in Sede.objects.all():
        if _normalizar_texto(sede.nombre) == objetivo:
            return sede
    return None


def _crear_perfil_importado(tipo, usuario, datos):
    """Crea el perfil específico (alumno/tutor) para un usuario recién creado.

    Devuelve None si todo va bien, o un string con el motivo del fallo de
    validación de negocio (p. ej. empresa inexistente) para reportarlo.
    """
    if tipo == 'alumno':
        carrera = _resolver_carrera(datos.get('carrera'))
        sede = _resolver_sede(datos.get('sede'))
        generacion = datos.get('generacion') or ''
        Alumno.objects.create(
            id=usuario,
            carrera=carrera,
            sede=sede,
            generacion=int(generacion) if generacion.isdigit() else None,
            matricula=datos.get('matricula') or None,
            created_at=timezone.now(),
            updated_at=timezone.now(),
        )
    elif tipo == 'tutor_udd':
        sede = _resolver_sede(datos.get('sede'))
        TutorUdd.objects.create(
            id=usuario,
            sede=sede,
            departamento=datos.get('departamento') or None,
            created_at=timezone.now(),
        )
    elif tipo == 'tutor_empresa':
        empresa = Empresa.objects.filter(nombre__icontains=datos['empresa']).first() if datos.get('empresa') else None
        if empresa is None:
            return f"empresa no encontrada ('{datos.get('empresa') or ''}')"
        TutorEmpresa.objects.create(
            id=usuario,
            empresa=empresa,
            cargo=datos.get('cargo') or None,
            created_at=timezone.now(),
        )
    return None


@require_roles()
def excel_import_view(request):
    import openpyxl
    from django.db import transaction

    # Período al que se asignarán los alumnos importados (respeta el Modo Auditoría).
    periodo, periodo_label = _periodo_contextual(request)
    results = None
    error = None

    rol_labels = {'alumno': 'alumno', 'tutor_udd': 'tutor UDD', 'tutor_empresa': 'tutor empresa'}

    if request.method == 'POST' and request.FILES.get('archivo'):
        archivo = request.FILES['archivo']
        tipo = (request.POST.get('tipo_usuario') or 'alumno').strip().lower()
        schema = IMPORT_SCHEMAS.get(tipo)

        if schema is None:
            error = 'Tipo de usuario no válido.'
        elif not archivo.name.lower().endswith('.xlsx'):
            error = 'El archivo debe tener formato .xlsx (Excel). Convierte el archivo y vuelve a intentarlo.'
        else:
            try:
                wb = openpyxl.load_workbook(archivo, data_only=True, read_only=True)
                ws = wb.active
            except Exception as exc:
                ws = None
                error = f'No se pudo abrir el archivo Excel: {exc}'

            if ws is not None:
                header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                if not header_row or all(c is None for c in header_row):
                    error = 'El archivo está vacío o no tiene una fila de encabezados en la primera línea.'
                else:
                    headers = [str(c or '').strip().lower() for c in header_row]
                    col_idx = {h: i for i, h in enumerate(headers) if h}
                    missing = [col for col in schema['required'] if col not in col_idx]
                    if missing:
                        error = f'Faltan columnas obligatorias para {rol_labels[tipo]}: {", ".join(missing)}.'
                        # Causa habitual del error: el listado de empresas se sube por el
                        # importador de personas. Se orienta al importador correcto.
                        messages.warning(
                            request,
                            "Si está intentando cargar el listado de empresas colaboradoras, por favor "
                            "utilice el botón 'Importar Empresas' ubicado en la pestaña de Empresas.",
                        )
                    else:
                        all_cols = schema['required'] + schema['optional']
                        created = skipped = errors_list = actualizados = 0
                        detail = []
                        # Mapa nombre normalizado → empresa, para guardar la empresa de
                        # referencia de cada alumno (no se crea proyecto, solo se anota).
                        empresas_norm = {}
                        if tipo == 'alumno':
                            for emp in Empresa.objects.all():
                                empresas_norm.setdefault(_normalizar_texto(emp.nombre), emp)

                        def _empresa_ref(datos):
                            return empresas_norm.get(_normalizar_texto(datos.get('empresa')))

                        # Período activo: los alumnos se enlazan a él vía AsignacionPeriodo.
                        periodo_asignacion = (
                            PeriodoAcademico.objects.filter(is_active=True).first()
                            if tipo == 'alumno' else None
                        )

                        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                            # Fila completamente vacía: se ignora silenciosamente.
                            if row is None or all(c is None or str(c).strip() == '' for c in row):
                                continue

                            def cell(name):
                                idx = col_idx.get(name)
                                if idx is None or idx >= len(row):
                                    return ''
                                value = row[idx]
                                return str(value).strip() if value is not None else ''

                            datos = {col: cell(col) for col in all_cols}
                            nombre = datos.get('nombre', '')
                            apellido = datos.get('apellido', '')
                            email = datos.get('email', '').lower()

                            # apellido no es obligatorio: si no viene, se separa del
                            # nombre completo (respeta apellidos compuestos).
                            if nombre and not apellido:
                                nombre, apellido = _separar_nombre_completo(nombre)
                            # Formato de nombre propio: inicial mayúscula, resto minúscula.
                            nombre = _titulo_es(nombre)
                            apellido = _titulo_es(apellido)
                            datos['nombre'] = nombre
                            datos['apellido'] = apellido

                            faltantes = [c for c in schema['required'] if not datos.get(c)]
                            if faltantes:
                                skipped += 1
                                detail.append({'row': row_num, 'status': 'skip', 'msg': f'Fila {row_num}: campos obligatorios vacíos ({", ".join(faltantes)})'})
                                continue

                            if Usuario.objects.filter(email=email).exists():
                                # Si el alumno ya existe, se actualiza su asignación de
                                # empresa en el período activo (backfill al reimportar).
                                emp_ref = _empresa_ref(datos) if tipo == 'alumno' else None
                                if emp_ref is not None and periodo_asignacion is not None:
                                    existente = Alumno.objects.filter(id__email=email).first()
                                    if existente is not None:
                                        AsignacionPeriodo.objects.update_or_create(
                                            alumno=existente, periodo=periodo_asignacion,
                                            defaults={'empresa': emp_ref, 'estado': 'activo'},
                                        )
                                        actualizados += 1
                                        detail.append({'row': row_num, 'status': 'skip', 'msg': f'Fila {row_num}: alumno existente, empresa actualizada'})
                                        continue
                                skipped += 1
                                detail.append({'row': row_num, 'status': 'skip', 'msg': f'Fila {row_num}: el correo ya existe ({email})'})
                                continue

                            try:
                                # Savepoint por fila: si una falla, se revierte solo
                                # esa fila y el resto del lote se conserva.
                                with transaction.atomic():
                                    usuario = Usuario.objects.create_user(
                                        email=email,
                                        password=None,
                                        rol=tipo,
                                        nombre=nombre,
                                        apellido=apellido,
                                        is_active=True,
                                        created_at=timezone.now(),
                                        updated_at=timezone.now(),
                                    )
                                    motivo = _crear_perfil_importado(tipo, usuario, datos)
                                    if motivo:
                                        raise ValueError(motivo)
                                    # No se crean proyectos: el alumno se vincula a su
                                    # empresa en el período activo vía AsignacionPeriodo.
                                    if tipo == 'alumno' and periodo_asignacion is not None:
                                        alumno_obj = Alumno.objects.get(pk=usuario.pk)
                                        AsignacionPeriodo.objects.update_or_create(
                                            alumno=alumno_obj, periodo=periodo_asignacion,
                                            defaults={'empresa': _empresa_ref(datos), 'estado': 'activo'},
                                        )
                            except Exception as exc:
                                errors_list += 1
                                detail.append({'row': row_num, 'status': 'error', 'msg': f'Fila {row_num}: {exc}'})
                            else:
                                created += 1
                                detail.append({'row': row_num, 'status': 'ok', 'msg': f'Fila {row_num}: {nombre} {apellido} ({email}) creado'})

                        # ── Feedback al administrador (mensajes de Django) ──
                        if created:
                            if tipo == 'alumno':
                                messages.success(
                                    request,
                                    f'{created} alumno(s) creado(s). Quedan sin asignar: los tutores '
                                    f'los integrarán a sus proyectos.'
                                )
                            else:
                                messages.success(request, f'{created} {rol_labels[tipo]}(s) creado(s) correctamente.')
                        if actualizados:
                            messages.info(request, f'{actualizados} alumno(s) ya existentes: se actualizó su empresa de referencia.')

                        results = {
                            'created': created,
                            'skipped': skipped,
                            'errors': errors_list,
                            'total': created + skipped + errors_list,
                            'detail': detail[:100],
                        }
            try:
                wb.close()
            except Exception:
                pass

    carreras = list(Carrera.objects.values_list('nombre', flat=True).order_by('nombre'))
    sedes_list = list(Sede.objects.values_list('nombre', flat=True).order_by('nombre'))
    context = {
        'periodo_label': periodo_label,
        'user_rol': _get_user_rol(request),
        'results': results,
        'error': error,
        'expected_cols': IMPORT_SCHEMAS['alumno']['required'] + IMPORT_SCHEMAS['alumno']['optional'],
        'carreras': carreras,
        'sedes': sedes_list,
    }
    return render(request, 'pages/excel_import.html', context)


# ─── BITÁCORA FILE UPLOAD ────────────────────────────────────────────────────

@require_roles('alumno', 'tutor_udd', 'tutor_empresa')
def bitacora_upload_view(request):
    if request.method != 'POST':
        return redirect('bitacora')

    pp_id = request.POST.get('proyecto_periodo_id')
    semana = request.POST.get('semana')
    archivo = request.FILES.get('archivo')
    alumno_id = request.POST.get('alumno_id', '')
    destino = f'/bitacora/?alumno={alumno_id}&sem={semana or 1}'

    if not (pp_id and semana and archivo):
        return redirect(destino)

    # Validación de extensión: solo PDF, PNG y JPG.
    ext = archivo.name.rsplit('.', 1)[-1].lower() if '.' in archivo.name else ''
    if ext not in ('pdf', 'png', 'jpg', 'jpeg'):
        messages.error(request, 'Solo se aceptan archivos PDF, PNG o JPG.')
        return redirect(destino)

    try:
        pp = ProyectoPeriodo.objects.get(pk=pp_id)
        semana_int = int(semana)
        bitacora, _ = Bitacora.objects.get_or_create(
            proyecto_periodo=pp,
            semana=semana_int,
            defaults={'created_at': timezone.now()},
        )
        # Se almacena el archivo real en la nueva tabla gestionada EvidenciaBitacora.
        EvidenciaBitacora.objects.create(bitacora=bitacora, archivo=archivo)
        messages.success(request, 'Evidencia subida correctamente.')
    except Exception as exc:
        messages.error(request, f'No se pudo subir la evidencia: {exc}')

    return redirect(destino)


# ─── CALIFICACIONES ──────────────────────────────────────────────────────────

@require_roles('tutor_udd', 'tutor_empresa')
def calificar_view(request):
    """Pantalla del tutor para evaluar las competencias de un hito por alumno.

    El tutor elige un alumno asignado y un hito; ve las competencias exigidas con
    un input por nota (1.0–7.0). Al guardar, se persiste cada nota en
    CalificacionCompetencia y se consolida la nota del hito (promedio ponderado
    por peso_pct) en EvaluacionHito.
    """
    user_rol = _get_user_rol(request)
    real_rol = getattr(request.user, 'rol', None)
    tutor_tipo = 'udd' if user_rol == 'tutor_udd' else 'empresa'
    # Aislamiento de período respetando el Modo Auditoría del administrador.
    periodo, periodo_label = _periodo_contextual(request)

    if not periodo:
        return render(request, 'pages/calificar.html', {
            'user_rol': user_rol, 'periodo_label': periodo_label, 'sin_datos': True,
            'mensaje': 'No hay un período activo para evaluar.',
        })

    pps = (
        ProyectoPeriodo.objects.filter(periodo=periodo, alumno__isnull=False)
        .select_related('proyecto__empresa', 'alumno__id')
    )
    if real_rol == 'tutor_udd':
        pps = pps.filter(tutores_udd=request.user.id)
    elif real_rol == 'tutor_empresa':
        pps = pps.filter(tutores_empresa=request.user.id)
    pps = list(pps.order_by('alumno__id__nombre', 'alumno__id__apellido'))

    alumno_uid = (request.GET.get('alumno') or request.POST.get('alumno') or '').strip()
    hito_id = (request.GET.get('hito') or request.POST.get('hito') or '').strip()

    current_pp = None
    if alumno_uid:
        current_pp = next((pp for pp in pps if str(pp.alumno.id_id) == alumno_uid), None)
    if current_pp is None and pps:
        current_pp = pps[0]

    # Hitos del período que evalúa este tipo de tutor (su tipo o 'ambos').
    hitos = [
        h for h in HitoEvaluacion.objects.filter(periodo=periodo).order_by('orden', 'semana')
        if (h.evaluador or 'ambos').strip().lower() in (tutor_tipo, 'ambos')
    ]
    current_hito = None
    if hito_id:
        current_hito = next((h for h in hitos if str(h.pk) == hito_id), None)
    if current_hito is None and hitos:
        current_hito = hitos[0]

    # POST: guardar notas por competencia.
    if request.method == 'POST' and current_pp and current_hito:
        competencias = list(CompetenciaHito.objects.filter(hito=current_hito).order_by('orden'))
        suma_pond = 0.0
        suma_pesos = 0
        guardadas = 0
        for comp in competencias:
            raw = (request.POST.get('comp_%s' % comp.pk) or '').strip().replace(',', '.')
            if not raw:
                continue
            try:
                val = float(raw)
            except ValueError:
                continue
            val = max(1.0, min(7.0, round(val, 1)))
            CalificacionCompetencia.objects.update_or_create(
                alumno=current_pp.alumno, competencia=comp,
                defaults={'nota': val, 'evaluado_por': request.user},
            )
            suma_pond += val * (comp.peso_pct or 0)
            suma_pesos += (comp.peso_pct or 0)
            guardadas += 1
        if guardadas and suma_pesos:
            nota_hito = round(suma_pond / suma_pesos, 1)
            EvaluacionHito.objects.update_or_create(
                hito=current_hito, proyecto_periodo=current_pp,
                defaults={
                    'evaluado_por': request.user,
                    'nota_calculada': nota_hito,
                    'fecha_evaluacion': timezone.now(),
                    'updated_at': timezone.now(),
                },
            )
            messages.success(request, 'Evaluación guardada. Nota del hito: %s.' % nota_hito)
        else:
            messages.error(request, 'Ingresa al menos una nota válida (1.0 a 7.0).')
        return redirect('%s?alumno=%s&hito=%s' % (request.path, current_pp.alumno.id_id, current_hito.pk))

    # GET: armar competencias con notas previas del alumno.
    competencias_data = []
    if current_hito and current_pp:
        previas = {
            c.competencia_id: c.nota
            for c in CalificacionCompetencia.objects.filter(
                alumno=current_pp.alumno, competencia__hito=current_hito
            )
        }
        for comp in CompetenciaHito.objects.filter(hito=current_hito).order_by('orden'):
            competencias_data.append({
                'id': comp.pk,
                'nombre': comp.nombre,
                'descripcion': comp.descripcion or '',
                'peso_pct': comp.peso_pct or 0,
                'nota': previas.get(comp.pk),
            })

    alumno_options = [
        {
            'id': str(pp.alumno.id_id),
            'label': f"{pp.alumno.id.nombre} {pp.alumno.id.apellido}".strip(),
            'empresa': pp.proyecto.empresa.nombre if pp.proyecto and pp.proyecto.empresa else '',
        }
        for pp in pps
    ]
    hito_options = [
        {'id': h.pk, 'nombre': h.nombre, 'semana': h.semana, 'peso_pct': h.peso_pct}
        for h in hitos
    ]

    student_name = ''
    student_company = ''
    if current_pp:
        student_name = f"{current_pp.alumno.id.nombre} {current_pp.alumno.id.apellido}".strip()
        student_company = current_pp.proyecto.empresa.nombre if current_pp.proyecto and current_pp.proyecto.empresa else 'Sin empresa'

    context = {
        'user_rol': user_rol,
        'periodo_label': periodo_label,
        'tutor_tipo': tutor_tipo,
        'alumno_options': alumno_options,
        'hito_options': hito_options,
        'selected_alumno_id': str(current_pp.alumno.id_id) if current_pp else '',
        'selected_hito_id': current_hito.pk if current_hito else '',
        'current_hito_nombre': current_hito.nombre if current_hito else '',
        'current_hito_semana': current_hito.semana if current_hito else '',
        'current_hito_peso': current_hito.peso_pct if current_hito else '',
        'competencias': competencias_data,
        'student_name': student_name,
        'student_company': student_company,
        'sin_datos': not (current_pp and current_hito and competencias_data),
    }
    return render(request, 'pages/calificar.html', context)


@require_roles('alumno')
def mis_notas_view(request):
    """Boletín de notas del alumno: nota automática de bitácoras + hitos + promedio.

    El promedio final solo se muestra como número cuando no quedan hitos
    pendientes de evaluación; en caso contrario se indica 'Promedio en cálculo'.
    """
    user_rol = _get_user_rol(request)
    real_rol = getattr(request.user, 'rol', None)
    # Aislamiento de período respetando el Modo Auditoría del administrador.
    periodo, periodo_label = _periodo_contextual(request)

    if not periodo:
        return render(request, 'pages/mis_notas.html', {
            'user_rol': user_rol, 'periodo_label': periodo_label, 'sin_datos': True,
            'mensaje': 'No hay un período activo.',
        })

    pp_qs = ProyectoPeriodo.objects.filter(periodo=periodo, alumno__isnull=False)
    if real_rol == 'alumno':
        pp_qs = pp_qs.filter(alumno_id=request.user.id)
    current_pp = pp_qs.select_related('proyecto__empresa', 'alumno__id').first()

    if current_pp is None:
        return render(request, 'pages/mis_notas.html', {
            'user_rol': user_rol, 'periodo_label': periodo_label, 'sin_datos': True,
            'mensaje': 'No tienes un proyecto asignado en el período activo.',
        })

    rows, promedio_final, en_calculo, pendientes = _calcular_boletin(periodo, current_pp)

    context = {
        'user_rol': user_rol,
        'periodo_label': periodo_label,
        'sin_datos': False,
        'student_name': f"{current_pp.alumno.id.nombre} {current_pp.alumno.id.apellido}".strip(),
        'student_company': current_pp.proyecto.empresa.nombre if current_pp.proyecto and current_pp.proyecto.empresa else 'Sin empresa',
        'project_name': current_pp.proyecto.titulo if current_pp.proyecto else 'Sin proyecto',
        'rows': rows,
        'promedio_final': promedio_final,
        'en_calculo': en_calculo,
        'pendientes': pendientes,
    }
    return render(request, 'pages/mis_notas.html', context)


# ─── PERFIL ──────────────────────────────────────────────────────────────────

def perfil_view(request):
    """Perfil del usuario autenticado. Muestra bloques distintos según `rol`."""
    user = request.user
    _, periodo_label = _periodo_activo()
    rol = user.rol
    user_rol = 'admin' if rol in ('admin', 'admin_iclae') else rol

    rol_labels = {
        'admin': 'Administrador',
        'admin_iclae': 'Administrador',
        'alumno': 'Alumno',
        'tutor_udd': 'Tutor UDD',
        'tutor_empresa': 'Tutor Empresa',
    }

    iniciales = f"{(user.nombre or '')[:1]}{(user.apellido or '')[:1]}".upper() or user.email[:2].upper()

    datos_rol = []
    enlaces_form = None
    if rol == 'alumno':
        alumno = Alumno.objects.select_related('carrera', 'sede').filter(pk=user.pk).first()
        if alumno is None:
            # Perfil mínimo para que el alumno pueda completar sus enlaces. carrera y
            # sede son obligatorias en la base, así que se asignan valores por defecto.
            carrera_def = Carrera.objects.order_by('pk').first()
            sede_def = Sede.objects.order_by('pk').first()
            if carrera_def and sede_def:
                alumno = Alumno.objects.create(
                    id=user, carrera=carrera_def, sede=sede_def,
                    created_at=timezone.now(), updated_at=timezone.now(),
                )
        if alumno:
            # El alumno edita sus tres enlaces de empleabilidad desde el perfil.
            if request.method == 'POST' and request.POST.get('action') == 'guardar_enlaces':
                enlaces_form = AlumnoEnlacesForm(request.POST, instance=alumno)
                if enlaces_form.is_valid():
                    enlaces_form.save()
                    messages.success(request, 'Tus enlaces se guardaron correctamente.')
                    return redirect('perfil')
                messages.error(request, 'Revisa los enlaces ingresados.')
            else:
                enlaces_form = AlumnoEnlacesForm(instance=alumno)
            datos_rol = [
                {'label': 'Carrera', 'value': alumno.carrera.nombre if alumno.carrera else ''},
                {'label': 'Sede', 'value': alumno.sede.nombre if alumno.sede else ''},
                {'label': 'Generación', 'value': alumno.generacion or ''},
            ]
    elif rol == 'tutor_empresa':
        tutor = TutorEmpresa.objects.select_related('empresa').filter(pk=user.pk).first()
        if tutor:
            datos_rol = [
                {'label': 'Empresa', 'value': tutor.empresa.nombre if tutor.empresa else ''},
                {'label': 'Cargo', 'value': tutor.cargo or ''},
            ]
    elif rol == 'tutor_udd':
        tutor = TutorUdd.objects.filter(pk=user.pk).first()
        if tutor:
            datos_rol = [
                {'label': 'Departamento', 'value': tutor.departamento or ''},
            ]

    context = {
        'periodo_label': periodo_label,
        'user_rol': user_rol,
        'perfil_nombre': user.get_full_name() or user.email,
        'perfil_email': user.email,
        'perfil_avatar_url': user.avatar_url,
        'perfil_iniciales': iniciales,
        'rol_label': rol_labels.get(rol, 'Usuario'),
        'datos_rol': datos_rol,
        'enlaces_form': enlaces_form,
    }
    return render(request, 'pages/perfil.html', context)


@require_roles('tutor_empresa', 'tutor_udd')
def asignar_badge_view(request):
    """Otorga una insignia del catálogo a un alumno; registra al tutor en otorgado_por."""
    periodo, periodo_label = _periodo_contextual(request)
    rol = getattr(request.user, 'rol', None)

    # Alumnos elegibles: los que el tutor acompaña en el período (admin ve todos los activos).
    if periodo is None:
        alumnos_qs = Alumno.objects.none()
    elif rol == 'tutor_empresa':
        alumnos_qs = Alumno.objects.filter(
            proyectoperiodo__periodo=periodo,
            proyectoperiodo__tutores_empresa__pk=request.user.pk,
        ).distinct()
    elif rol == 'tutor_udd':
        alumnos_qs = Alumno.objects.filter(
            proyectoperiodo__periodo=periodo,
            proyectoperiodo__tutores_udd__pk=request.user.pk,
        ).distinct()
    else:  # admin
        alumnos_qs = Alumno.objects.filter(proyectoperiodo__periodo=periodo).distinct()
    alumnos_qs = alumnos_qs.select_related('id').order_by('id__nombre', 'id__apellido')

    if request.method == 'POST':
        form = AsignacionBadgeForm(request.POST, alumnos_qs=alumnos_qs)
        if form.is_valid():
            alumno = form.cleaned_data['alumno']
            badge = form.cleaned_data['badge']
            # unique_together (alumno, badge, periodo): get_or_create evita duplicar.
            _, creado = AlumnoBadge.objects.get_or_create(
                alumno=alumno,
                badge=badge,
                periodo=periodo,
                defaults={
                    'motivo': form.cleaned_data.get('motivo') or None,
                    'otorgado_por': request.user,
                    'fecha_logro': timezone.now(),
                },
            )
            if creado:
                messages.success(request, 'Insignia otorgada a %s.' % alumno.id.get_full_name())
            else:
                messages.info(request, 'Ese alumno ya tenía esa insignia en el período actual.')
            return redirect('asignar_badge')
        messages.error(request, 'No se pudo otorgar la insignia. Revisa los datos ingresados.')
    else:
        form = AsignacionBadgeForm(alumnos_qs=alumnos_qs)

    context = {
        'periodo_label': periodo_label,
        'user_rol': 'admin' if rol == 'admin' else rol,
        'form': form,
        'badges_recientes': (
            AlumnoBadge.objects.filter(otorgado_por=request.user)
            .select_related('alumno__id', 'badge')
            .order_by('-fecha_logro')[:10]
        ),
    }
    return render(request, 'pages/asignar_badge.html', context)


# ─── EXPORTAR ────────────────────────────────────────────────────────────────

def _xlsx_response(filename, sheet_title, headers, rows):
    """Construye un archivo Excel (.xlsx) real con encabezados en negrita.

    headers: lista de títulos de columna. rows: iterable de listas (una por fila).
    Devuelve un HttpResponse listo para descargar.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31] or 'Hoja1'

    ws.append(list(headers))
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical='center')
    ws.freeze_panes = 'A2'

    for row in rows:
        ws.append(list(row))

    # Ancho de columnas según el contenido más largo (acotado).
    for idx, header in enumerate(headers, start=1):
        largest = len(str(header))
        for row in rows:
            if idx <= len(row) and row[idx - 1] is not None:
                largest = max(largest, len(str(row[idx - 1])))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(largest + 2, 12), 48)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@require_roles()
def exportar_alumnos_view(request):
    from django.db.models import Max as _Max2

    q = (request.GET.get('q') or '').strip()
    carrera_filter = (request.GET.get('carrera') or '').strip().lower()
    sede_filter = (request.GET.get('sede') or '').strip().lower()
    # Aislamiento de período respetando el Modo Auditoría del administrador.
    periodo, _ = _periodo_contextual(request)

    qs = (
        Alumno.objects.select_related('id', 'carrera', 'sede')
        .annotate(empresa_nombre=_Max2(
            'proyectoperiodo__proyecto__empresa__nombre',
            filter=Q(proyectoperiodo__periodo=periodo),
        ))
        .order_by('id__nombre', 'id__apellido')
    )
    if q:
        qs = qs.filter(
            Q(id__nombre__icontains=q) | Q(id__apellido__icontains=q) | Q(id__email__icontains=q)
        )
    if carrera_filter:
        qs = qs.filter(carrera__nombre__icontains=carrera_filter)
    if sede_filter:
        qs = qs.filter(sede__nombre__icontains=sede_filter)

    headers = ['Nombre', 'Email', 'Carrera', 'Sede', 'Generación', 'Empresa']
    rows = [
        [
            f"{a.id.nombre} {a.id.apellido}".strip(),
            a.id.email,
            a.carrera.nombre if a.carrera else '',
            a.sede.nombre if a.sede else '',
            a.generacion or '',
            a.empresa_nombre or '',
        ]
        for a in qs
    ]
    return _xlsx_response('alumnos.xlsx', 'Alumnos', headers, rows)


# ─── GESTIONAR USUARIOS ───────────────────────────────────────────────────────

def _proyecto_para_empresa(empresa, creador):
    """Devuelve un Proyecto de la empresa para enlazar al alumno (Opción B).

    El modelo intermedio ProyectoPeriodo exige un Proyecto, así que se reutiliza
    uno existente de la empresa (preferentemente activo) y, solo si la empresa no
    tiene ninguno, se crea uno mínimo para soportar la asignación. Así el enlace
    alumno ↔ empresa ↔ período nunca falla por ausencia de proyecto.
    """
    proyecto = (
        Proyecto.objects.filter(empresa=empresa, is_active=True).order_by('-created_at').first()
        or Proyecto.objects.filter(empresa=empresa).order_by('-created_at').first()
    )
    if proyecto is None:
        proyecto = Proyecto.objects.create(
            empresa=empresa,
            titulo=f'Asignación {empresa.nombre}'[:255],
            is_active=True,
            created_by=creador,
            created_at=timezone.now(),
            updated_at=timezone.now(),
        )
    return proyecto


@require_roles()
def gestionar_usuarios_view(request):
    user_rol = _get_user_rol(request)
    if user_rol != 'admin':
        return redirect('/')

    _, periodo_label = _periodo_activo()
    saved_ok = False
    error = None

    if request.method == 'POST':
        action = request.POST.get('action', '')

        if action == 'crear_usuario':
            form = UsuarioCreacionForm(request.POST)
            if not form.is_valid():
                # Validación fallida (incluye empresa obligatoria para tutor empresa):
                # se muestra el mensaje amigable y no se toca la base de datos.
                error = ' '.join(e for errores in form.errors.values() for e in errores)
            else:
                cd = form.cleaned_data
                rol = cd['rol']
                # El enlace del alumno con su empresa usa el período contextual
                # (auditoría en sesión o período global activo).
                periodo_enlace, _ = _periodo_contextual(request)
                if rol == 'alumno' and periodo_enlace is None:
                    messages.error(
                        request,
                        'No hay un período activo para enlazar al alumno con su empresa. '
                        'Activa un período e inténtalo de nuevo.'
                    )
                    return redirect('gestionar_usuarios')
                try:
                    # Todo o nada: si la creación del perfil específico o del enlace
                    # falla, el usuario base se revierte y no queda nada huérfano.
                    with transaction.atomic():
                        nuevo = Usuario.objects.create_user(
                            email=cd['email'],
                            password=cd['password'],
                            rol=rol,
                            nombre=cd['nombre'],
                            apellido=cd['apellido'],
                            is_active=True,
                            created_at=timezone.now(),
                            updated_at=timezone.now(),
                        )
                        if rol == 'alumno':
                            carrera_obj = Carrera.objects.filter(id=cd['carrera_id']).first() if cd['carrera_id'] else None
                            sede_obj = Sede.objects.filter(id=cd['sede_id']).first() if cd['sede_id'] else None
                            alumno = Alumno.objects.create(
                                id=nuevo,
                                carrera=carrera_obj,
                                sede=sede_obj,
                                created_at=timezone.now(),
                                updated_at=timezone.now(),
                            )
                            # Opción B: enlace alumno ↔ empresa ↔ período mediante el
                            # modelo intermedio ProyectoPeriodo, en la misma transacción.
                            empresa_enlace = cd['empresa_selector']
                            proyecto_enlace = _proyecto_para_empresa(empresa_enlace, nuevo)
                            ProyectoPeriodo.objects.create(
                                proyecto=proyecto_enlace,
                                periodo=periodo_enlace,
                                alumno=alumno,
                                sede=sede_obj,
                                created_at=timezone.now(),
                                updated_at=timezone.now(),
                            )
                        elif rol == 'tutor_udd':
                            sede_obj = Sede.objects.filter(id=cd['sede_id']).first() if cd['sede_id'] else None
                            TutorUdd.objects.create(
                                id=nuevo,
                                sede=sede_obj,
                                departamento=cd['departamento'] or None,
                                created_at=timezone.now(),
                            )
                        elif rol == 'tutor_empresa':
                            empresa_obj = Empresa.objects.filter(id=cd['empresa_id']).first()
                            TutorEmpresa.objects.create(
                                id=nuevo,
                                empresa=empresa_obj,
                                cargo=cd['cargo'] or None,
                                created_at=timezone.now(),
                            )
                    saved_ok = True
                except IntegrityError:
                    # La transacción ya hizo rollback: la plataforma sigue estable.
                    messages.error(
                        request,
                        'No se pudo crear el usuario: faltan datos obligatorios del perfil o se '
                        'produjo un conflicto de integridad. No se guardó ningún registro.'
                    )
                    return redirect('gestionar_usuarios')

        elif action == 'cambiar_password':
            usuario_id = request.POST.get('usuario_id', '').strip()
            new_password = request.POST.get('new_password', '').strip()
            if usuario_id and new_password:
                try:
                    u = Usuario.objects.get(id=usuario_id)
                    u.set_password(new_password)
                    u.updated_at = timezone.now()
                    u.save(update_fields=['password', 'updated_at'])
                    saved_ok = True
                except Exception as exc:
                    error = f'Error: {exc}'

        elif action == 'toggle_activo':
            usuario_id = request.POST.get('usuario_id', '').strip()
            if usuario_id:
                try:
                    u = Usuario.objects.get(id=usuario_id)
                    u.is_active = not u.is_active
                    u.updated_at = timezone.now()
                    u.save(update_fields=['is_active', 'updated_at'])
                    saved_ok = True
                except Exception as exc:
                    error = f'Error: {exc}'

    rol_filter = (request.GET.get('rol_f') or '').strip().lower()
    q = (request.GET.get('q') or '').strip()

    usuarios_qs = Usuario.objects.order_by('nombre', 'apellido')
    if rol_filter and rol_filter in ('admin', 'alumno', 'tutor_udd', 'tutor_empresa'):
        usuarios_qs = usuarios_qs.filter(rol=rol_filter)
    if q:
        usuarios_qs = usuarios_qs.filter(
            Q(nombre__icontains=q) | Q(apellido__icontains=q) | Q(email__icontains=q)
        )

    usuarios = [
        {
            'id': str(u.id),
            'nombre': f"{u.nombre} {u.apellido}".strip(),
            'email': u.email,
            'rol': u.rol,
            'has_password': u.has_usable_password(),
            'is_active': u.is_active,
            'created_at': u.created_at,
        }
        for u in usuarios_qs[:200]
    ]

    carreras = list(Carrera.objects.values('id', 'nombre').order_by('nombre'))
    sedes_list = list(Sede.objects.values('id', 'nombre').order_by('nombre'))
    empresas_list = list(Empresa.objects.values('id', 'nombre').order_by('nombre'))

    context = {
        'periodo_label': periodo_label,
        'user_rol': user_rol,
        'usuarios': usuarios,
        'usuarios_count': len(usuarios),
        'saved_ok': saved_ok,
        'error': error,
        'filters': {'q': q, 'rol': rol_filter},
        'carreras': carreras,
        'sedes': sedes_list,
        'empresas': empresas_list,
    }
    return render(request, 'pages/gestionar_usuarios.html', context)


# ─── EXPORTAR ────────────────────────────────────────────────────────────────

@require_roles()
def exportar_empresas_view(request):
    q = (request.GET.get('q') or '').strip()
    rubro_filter = (request.GET.get('rubro') or '').strip().lower()
    # Aislamiento de período respetando el Modo Auditoría del administrador.
    periodo, _ = _periodo_contextual(request)

    qs = (
        Empresa.objects.annotate(
            practicantes_count=Count('proyecto__proyectoperiodo', filter=Q(proyecto__proyectoperiodo__alumno__isnull=False, proyecto__proyectoperiodo__periodo=periodo), distinct=True),
            proyectos_count=Count('proyecto', filter=Q(proyecto__proyectoperiodo__periodo=periodo), distinct=True),
        ).order_by('nombre')
    )
    if q:
        qs = qs.filter(Q(nombre__icontains=q) | Q(rubro__icontains=q))
    if rubro_filter:
        qs = qs.filter(rubro__icontains=rubro_filter)

    headers = ['Empresa', 'Rubro', 'Presencia', 'eNPS', 'Alumnos', 'Proyectos', 'Contacto', 'Email contacto']
    rows = [
        [
            e.nombre or '',
            e.rubro or '',
            e.presencia or '',
            e.enps_score or '',
            e.practicantes_count,
            e.proyectos_count,
            e.contacto_nombre or '',
            e.contacto_email or '',
        ]
        for e in qs
    ]
    return _xlsx_response('empresas.xlsx', 'Empresas', headers, rows)


# Carga progresiva de empresas. Lo único imprescindible es el nombre; el resto puede
# quedar en blanco (se guarda como nulo) y completarse después. En la plantilla, las
# opcionales llevan el sufijo "(opcional)" en el encabezado; la importación lo normaliza.
EMPRESA_IMPORT_OBLIGATORIAS = ['nombre']
EMPRESA_IMPORT_OPCIONALES = ['rubro', 'ubicacion', 'contacto_nombre', 'contacto_email', 'presencia', 'descripcion']
# Alias tolerados para la columna del nombre (se comparan sin acentos ni espacios).
EMPRESA_IMPORT_NOMBRE_ALIASES = ['nombre', 'empresa', 'nombre empresa', 'nombre de la empresa', 'organizacion', 'razon social']


def empresas_plantilla_view(request):
    """Descarga una plantilla Excel (.xlsx) vacía con los encabezados de Empresa.

    La primera fila trae el único encabezado obligatorio (nombre) y el resto rotulados
    con el sufijo (opcional): rubro, ubicacion, contacto_nombre, contacto_email,
    presencia y descripcion. La importación normaliza el sufijo al leer.
    """
    encabezados = list(EMPRESA_IMPORT_OBLIGATORIAS) + [f'{col} (opcional)' for col in EMPRESA_IMPORT_OPCIONALES]
    return _xlsx_response('plantilla_empresas.xlsx', 'Empresas', encabezados, [])


def empresas_import_view(request):
    """Importación masiva de empresas desde un archivo Excel (.xlsx).

    Lectura ultra tolerante: si el archivo trae una sola columna, esa columna se toma
    como nombre aunque la primera fila sea ya un dato y no un encabezado. El único
    criterio para descartar una fila es que el nombre venga vacío o nulo. Inserta con
    get_or_create por nombre saneado (idempotente) y cierra con un resumen exacto.
    """
    import openpyxl

    user_rol = _get_user_rol(request)
    _, periodo_label = _periodo_activo()

    if request.method == 'POST':
        if getattr(request.user, 'rol', None) != 'admin':
            return HttpResponseForbidden('No tienes permisos para importar empresas.')

        archivo = request.FILES.get('archivo')
        if archivo is None:
            messages.error(request, 'Selecciona un archivo Excel (.xlsx) para importar.')
            return redirect('empresas_import')
        if not archivo.name.lower().endswith('.xlsx'):
            messages.error(request, 'El archivo debe tener formato Excel (.xlsx). Sube el archivo original sin convertirlo.')
            return redirect('empresas_import')

        try:
            wb = openpyxl.load_workbook(archivo, data_only=True, read_only=True)
            ws = wb.active
        except Exception:
            messages.error(request, 'No se pudo leer el archivo: parece dañado o no es un Excel válido. Revísalo e inténtalo de nuevo.')
            return redirect('empresas_import')

        # ── Lectura ultra tolerante: se cargan todas las filas con algún contenido ──
        filas = [
            fila for fila in ws.iter_rows(min_row=1, values_only=True)
            if fila is not None and not all(c is None or str(c).strip() == '' for c in fila)
        ]
        if not filas:
            messages.error(request, 'El archivo está vacío: no tiene datos para importar.')
            return redirect('empresas_import')

        # Encabezado tentativo (primera fila), normalizado y sin el sufijo "(opcional)".
        encabezado = [str(c or '').strip().lower().replace('(opcional)', '').strip() for c in filas[0]]
        col_idx = {h: i for i, h in enumerate(encabezado) if h}

        # Mapeo flexible del nombre entre alias frecuentes (comparando sin acentos).
        nombre_idx = col_idx.get('nombre')
        if nombre_idx is None:
            headers_norm = {_normalizar_texto(h): i for h, i in col_idx.items()}
            for alias in EMPRESA_IMPORT_NOMBRE_ALIASES:
                idx = headers_norm.get(_normalizar_texto(alias))
                if idx is not None:
                    nombre_idx = idx
                    break

        # ¿La primera fila es realmente un encabezado? Lo es si reconocemos la columna
        # del nombre o cualquier columna conocida de Empresa. Si no reconocemos nada,
        # se asume que la primera columna es el nombre y que la primera fila ya es un
        # dato: así un Excel de una sola columna (incluso de 1x1, sin cabecera) entra.
        columnas_conocidas = set(EMPRESA_IMPORT_OPCIONALES) | {'nombre'}
        hay_encabezado = nombre_idx is not None or any(h in columnas_conocidas for h in col_idx)
        if nombre_idx is None:
            nombre_idx = 0  # única o primera columna = nombre

        filas_datos = filas[1:] if hay_encabezado else filas
        # Los campos opcionales solo se mapean cuando hubo una cabecera reconocible.
        usar_opcionales = hay_encabezado
        presencias_validas = ('chile', 'multinacional')

        def celda(fila, idx):
            if idx is None or idx >= len(fila):
                return ''
            valor = fila[idx]
            return str(valor).strip() if valor is not None else ''

        def limpiar_opcional(valor):
            """Cadena vacía, solo espacios o nan -> None; si hay valor real, sin espacios."""
            texto = (valor or '').strip()
            if not texto or texto.lower() == 'nan':
                return None
            return texto

        def opcional(fila, name, limite):
            """Opcional saneado y truncado; None si no hubo cabecera o viene vacío."""
            if not usar_opcionales:
                return None
            texto = limpiar_opcional(celda(fila, col_idx.get(name)))
            return texto[:limite] if texto else None

        # Nombres existentes normalizados (sin acentos ni mayúsculas) para no duplicar.
        nombres_existentes = {_normalizar_texto(n) for n in Empresa.objects.values_list('nombre', flat=True) if n}
        nombres_vistos = set()

        ahora = timezone.now()
        creadas = 0
        omitidas_duplicadas = 0
        omitidas_sin_nombre = 0
        fallidas = 0

        for fila in filas_datos:
            # ── Único criterio de descarte: nombre vacío o nulo (NaN) ──
            nombre = limpiar_opcional(celda(fila, nombre_idx))
            if not nombre:
                omitidas_sin_nombre += 1
                continue

            # Formato de razón social: Mayúscula inicial, conservando siglas (S.A., SpA).
            nombre = _titulo_es(nombre, siglas=True)
            clave = _normalizar_texto(nombre)
            if clave in nombres_existentes or clave in nombres_vistos:
                omitidas_duplicadas += 1
                continue
            nombres_vistos.add(clave)

            # presencia es un enum de PostgreSQL: solo acepta None o etiqueta válida.
            presencia = opcional(fila, 'presencia', 50)
            if presencia is not None:
                presencia = presencia.lower()
                if presencia not in presencias_validas:
                    presencia = None

            # get_or_create por nombre saneado; cada fila en su savepoint para que un
            # error aislado no tumbe el resto del lote.
            try:
                with transaction.atomic():
                    _, creado = Empresa.objects.get_or_create(
                        nombre=nombre[:255],
                        defaults={
                            'rubro': opcional(fila, 'rubro', 150),
                            'ubicacion': opcional(fila, 'ubicacion', 255),
                            'presencia': presencia,
                            'descripcion': opcional(fila, 'descripcion', None),
                            'contacto_nombre': opcional(fila, 'contacto_nombre', 150),
                            'contacto_email': opcional(fila, 'contacto_email', 254),
                            'is_active': True,
                            'created_at': ahora,
                            'updated_at': ahora,
                        },
                    )
                if creado:
                    creadas += 1
                else:
                    omitidas_duplicadas += 1
            except Exception:
                fallidas += 1

        # Resumen exacto (conteo correcto aunque sea una sola empresa).
        partes = ['%d empresa(s) creada(s)' % creadas]
        if omitidas_duplicadas:
            partes.append('%d omitida(s) por duplicidad' % omitidas_duplicadas)
        if omitidas_sin_nombre:
            partes.append('%d omitida(s) sin nombre' % omitidas_sin_nombre)
        if fallidas:
            partes.append('%d con error de datos' % fallidas)
        messages.success(request, 'Importación finalizada. ' + ', '.join(partes) + '.')
        return redirect('empresas')

    return render(request, 'pages/empresas_import.html', {
        'user_rol': user_rol,
        'periodo_label': periodo_label,
    })


@require_roles()
def descargar_plantilla_view(request):
    """Descarga una plantilla Excel (.xlsx) vacía con los encabezados exactos que
    espera la importación, lista para poblar y volver a subir sin problemas de
    formato. El tipo se elige con el parámetro tipo (alumno, tutor_udd, tutor_empresa).
    """
    tipo = (request.GET.get('tipo') or 'alumno').strip()
    schema = IMPORT_SCHEMAS.get(tipo)
    if not schema:
        tipo = 'alumno'
        schema = IMPORT_SCHEMAS['alumno']
    columnas = list(schema['required']) + list(schema['optional'])
    return _xlsx_response(f'plantilla_{tipo}.xlsx', 'Plantilla', columnas, [])
